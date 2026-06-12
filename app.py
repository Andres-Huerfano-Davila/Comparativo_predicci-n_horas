# -*- coding: utf-8 -*-
"""
Comparativo y predicción de horas de nómina - V14
Creado para comparar CCNómina pagada (mes vencido) vs provisión vs proyección,
con homologación basada en Maestro Posición -> Función -> Cargo homologado.
"""

import io
import os
import re
import math
import zipfile
import hashlib
import unicodedata
from datetime import datetime
from functools import reduce
from typing import Dict, List, Tuple, Optional, Any

import numpy as np
import pandas as pd
import streamlit as st

# ==============================
# Configuración general
# ==============================
st.set_page_config(
    page_title="Comparativo y predicción de horas de nómina",
    page_icon="🦜",
    layout="wide",
)

APP_VERSION = "V14.6 - Descargas seguras + Excel ejecutivo liviano + HC por periodo"
ORANGE = "#F26A21"
BLUE = "#005AA9"
GREEN = "#2E8B57"
RED = "#D62828"
YELLOW = "#F4B400"
GRAY_BG = "#F4F6F8"
DARK = "#333333"

CONCEPTOS = {
    "Y220": "Rec. Noc.",
    "Y221": "Rec. Dom noc",
    "Y300": "Hora Extra",
    "Y305": "Hora Extra",
    "Y310": "Hora Extra",
    "Y315": "Hora Extra",
    "Y350": "Compensatorio",
    "YM01": "Rec. Dom",
}
CONCEPTOS_SET = set(CONCEPTOS.keys())

INTERFAZ_MAP = {
    "Y540": "Y220",
    "Y541": "Y221",
    "Y542": "Y300",
    "Y543": "Y305",
    "Y544": "Y310",
    "Y545": "Y315",
    "Y546": "Y350",
    "Y547": "YM01",
}

KEY_DETAIL = ["periodo_novedad", "area_negocio", "cargo_homologado", "ceco", "concepto", "tipo_hora"]
KEY_EXEC = ["periodo_novedad", "area_negocio", "cargo_homologado", "concepto", "tipo_hora"]
KEY_HC = ["periodo_novedad", "area_negocio", "cargo_homologado"]
MAX_SCREEN_ROWS = 5000

# ==============================
# CSS / Branding
# ==============================
st.markdown(
    f"""
    <style>
    .main .block-container {{ padding-top: 1.0rem; }}
    .jmc-header {{
        background: linear-gradient(90deg, {ORANGE} 0%, #FF8A3D 100%);
        padding: 18px 22px;
        border-radius: 18px;
        color: white;
        box-shadow: 0 6px 18px rgba(0,0,0,0.08);
        margin-bottom: 15px;
    }}
    .jmc-title {{ font-size: 29px; font-weight: 800; margin: 0; }}
    .jmc-subtitle {{ font-size: 14px; opacity: .95; margin-top: 4px; }}
    .small-note {{ color: #6b7280; font-size: 13px; }}
    .stButton button {{ border-radius: 10px; font-weight: 700; }}
    div[data-testid="stMetric"] {{
        background: white;
        border: 1px solid #e5e7eb;
        border-radius: 16px;
        padding: 12px;
        box-shadow: 0 2px 10px rgba(0,0,0,0.04);
    }}
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    f"""
    <div class="jmc-header">
      <div class="jmc-title">🦜 Comparativo y predicción de horas de nómina</div>
      <div class="jmc-subtitle">{APP_VERSION} · Pagado real mes vencido · Provisión · Proyección · Headcount · Homologación por función</div>
    </div>
    """,
    unsafe_allow_html=True,
)

# ==============================
# Utilidades base
# ==============================
def strip_accents(text: Any) -> str:
    if pd.isna(text):
        return ""
    text = str(text)
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def norm_key(text: Any) -> str:
    s = strip_accents(text).upper().strip()
    s = re.sub(r"[\n\r\t]+", " ", s)
    s = re.sub(r"[^A-Z0-9]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def clean_text(text: Any) -> str:
    if pd.isna(text):
        return ""
    return str(text).strip()


def clean_code(value: Any) -> str:
    if pd.isna(value):
        return ""
    s = str(value).strip()
    if s.endswith(".0") and re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    s = re.sub(r"\s+", "", s)
    return s


def clean_sap(value: Any) -> str:
    s = clean_code(value)
    if not s or norm_key(s) in {"ERROR", "NAN", "NONE", "NULL"}:
        return ""
    digits = re.sub(r"\D", "", s)
    return digits if len(digits) >= 5 else ""


def clean_concept(value: Any) -> str:
    s = clean_code(value).upper()
    return s


def parse_number(x: Any) -> float:
    if pd.isna(x):
        return 0.0
    if isinstance(x, (int, float, np.integer, np.floating)):
        if math.isnan(float(x)):
            return 0.0
        return float(x)
    s = str(x).strip()
    if not s or norm_key(s) in {"NAN", "NONE", "NULL", "-"}:
        return 0.0
    s = s.replace("COP", "").replace("$", "").replace(" ", "")
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1]
    # europeo: 1.234.567,89
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    elif "." in s:
        # Si parece separador de miles: 55.713 o 1.234.567
        if re.fullmatch(r"-?\d{1,3}(\.\d{3})+", s):
            s = s.replace(".", "")
    try:
        val = float(s)
        return -val if neg else val
    except Exception:
        return 0.0


def format_money(x: Any) -> str:
    try:
        v = float(x)
    except Exception:
        v = 0.0
    return "$ " + f"{v:,.0f}".replace(",", ".")


def format_qty(x: Any) -> str:
    try:
        v = float(x)
    except Exception:
        v = 0.0
    s = f"{v:,.2f}"
    return s.replace(",", "X").replace(".", ",").replace("X", ".")


def format_int(x: Any) -> str:
    try:
        v = float(x)
    except Exception:
        v = 0.0
    return f"{v:,.0f}".replace(",", ".")


def format_pct(x: Any) -> str:
    try:
        v = float(x)
    except Exception:
        v = 0.0
    s = f"{v:,.2f}%"
    return s.replace(",", "X").replace(".", ",").replace("X", ".")


def parse_period_any(value: Any, fallback_name: str = "") -> str:
    """Retorna MM.YYYY."""
    if pd.notna(value) and not isinstance(value, (bytes, bytearray)):
        if isinstance(value, (pd.Timestamp, datetime)):
            return f"{value.month:02d}.{value.year}"
        s = str(value).strip()
        if s and norm_key(s) not in {"NAN", "NONE", "NULL", "ERROR"}:
            # 202602
            m = re.search(r"(20\d{2})(0[1-9]|1[0-2])", s)
            if m:
                return f"{m.group(2)}.{m.group(1)}"
            # 02.2026 o 02/2026 o 02-2026
            m = re.search(r"(0?[1-9]|1[0-2])[\.\-/ ](20\d{2})", s)
            if m:
                return f"{int(m.group(1)):02d}.{m.group(2)}"
            # 022026
            m = re.search(r"\b(0[1-9]|1[0-2])(20\d{2})\b", s)
            if m:
                return f"{m.group(1)}.{m.group(2)}"
    if fallback_name:
        return parse_period_any(str(fallback_name), "")
    return ""


def normalize_period_value(value: Any, fallback_name: str = "") -> str:
    """Normaliza cualquier período a MM.YYYY. Corrige casos como 1.2026 -> 01.2026."""
    if pd.isna(value):
        value = ""
    s = str(value).strip()
    if s in {"", "nan", "NaN", "None", "NULL"}:
        s = ""
    # Caso en que pandas leyó 01.2026 como float 1.2026
    m = re.fullmatch(r"([1-9]|1[0-2])\.(20\d{2})", s)
    if m:
        return f"{int(m.group(1)):02d}.{m.group(2)}"
    p = parse_period_any(s, fallback_name)
    if p:
        return p
    # Último recurso: detectar un bloque de 6 dígitos en el nombre/valor
    txt = f"{s} {fallback_name}"
    for m in re.finditer(r"(?<!\d)(0[1-9]|1[0-2])(20\d{2})(?!\d)", txt):
        return f"{m.group(1)}.{m.group(2)}"
    for m in re.finditer(r"(?<!\d)(20\d{2})(0[1-9]|1[0-2])(?!\d)", txt):
        return f"{m.group(2)}.{m.group(1)}"
    return ""


def prev_period(period: str) -> str:
    period = normalize_period_value(period)
    if not period:
        return ""
    m = re.match(r"(0[1-9]|1[0-2])\.(20\d{2})", str(period))
    if not m:
        return ""
    month = int(m.group(1)); year = int(m.group(2))
    month -= 1
    if month == 0:
        month = 12; year -= 1
    return f"{month:02d}.{year}"


def period_sort_key(period: str) -> int:
    m = re.match(r"(0[1-9]|1[0-2])\.(20\d{2})", str(period))
    if not m:
        return 999999
    return int(m.group(2))*100 + int(m.group(1))


def find_col(df: pd.DataFrame, candidates: List[str], required: bool = False) -> Optional[str]:
    normalized_cols = {norm_key(c): c for c in df.columns}
    cand_norm = [norm_key(c) for c in candidates]
    for cn in cand_norm:
        if cn in normalized_cols:
            return normalized_cols[cn]
    # fuzzy contains, prefer shortest
    for cn in cand_norm:
        matches = [orig for nk, orig in normalized_cols.items() if cn and (cn in nk or nk in cn)]
        if matches:
            return sorted(matches, key=lambda x: len(str(x)))[0]
    if required:
        raise ValueError(f"No encontré columna para: {candidates}")
    return None


def coalesce_cols(df: pd.DataFrame, cols: List[Optional[str]]) -> pd.Series:
    out = pd.Series([""] * len(df), index=df.index, dtype="object")
    for c in cols:
        if c and c in df.columns:
            s = df[c].apply(clean_text)
            out = out.mask(out.eq(""), s)
    return out


def drop_duplicated_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    return df.loc[:, ~pd.Index(df.columns).duplicated()].copy()


def classify_area(ceco: Any = "", division: Any = "", tipo: Any = "", area_nomina: Any = "", cargo: Any = "") -> str:
    ceco_s = clean_code(ceco)
    combo = norm_key(" ".join([clean_text(division), clean_text(tipo), clean_text(area_nomina), clean_text(cargo)]))
    if "BODEGA CANASTO" in combo or re.search(r"\bBDC\b", combo):
        return "BDC"
    if "LOGISTICA" in combo or "DISTR CENTER" in combo or "CENTRO DE DISTRIB" in combo or "CEDI" in combo:
        return "CEDI"
    if "ADMINISTRAT" in combo or "OFICINA" in combo or "SOPORTE" in combo or "HQ" in combo:
        return "Oficina Soporte"
    if ceco_s.startswith("102"):
        return "CEDI"
    if ceco_s.startswith("103"):
        return "Oficina Soporte"
    if ceco_s.startswith("101"):
        return "Tiendas"
    return "Sin clasificar"


def is_manager_excl(area_personal: Any = "", area_nomina: Any = "", cargo: Any = "") -> bool:
    s = norm_key(" ".join([clean_text(area_personal), clean_text(area_nomina), clean_text(cargo)]))
    # No excluir Non Manager
    if "NON MANAGER" in s:
        return False
    # Manager I, II, III, IV o Manager 1-4
    if re.search(r"\bMANAGER\s*(I|II|III|IV|1|2|3|4)\b", s):
        return True
    return False


def ensure_key_types(df: pd.DataFrame, keys: List[str]) -> pd.DataFrame:
    df = df.copy()
    for k in keys:
        if k not in df.columns:
            df[k] = ""
        df[k] = df[k].fillna("").astype(str)
        if k.startswith("periodo"):
            df[k] = df[k].apply(normalize_period_value)
    return df

# ==============================
# Lectura de archivos
# ==============================
def get_bytes(uploaded_file) -> bytes:
    if uploaded_file is None:
        return b""
    return uploaded_file.getvalue()


def _read_excel_bytes(data: bytes, name: str, sheet_name=0, nrows=None, usecols=None) -> pd.DataFrame:
    ext = os.path.splitext(name)[1].lower()
    bio = io.BytesIO(data)
    # xlsb
    if ext == ".xlsb":
        return pd.read_excel(bio, sheet_name=sheet_name, engine="pyxlsb", nrows=nrows, usecols=usecols)
    # xlsx / xlsm / xls
    for engine in ["calamine", "openpyxl", None]:
        try:
            bio.seek(0)
            if engine:
                return pd.read_excel(bio, sheet_name=sheet_name, engine=engine, nrows=nrows, usecols=usecols)
            return pd.read_excel(bio, sheet_name=sheet_name, nrows=nrows, usecols=usecols)
        except Exception:
            continue
    # último intento genérico
    bio.seek(0)
    return pd.read_excel(bio, sheet_name=sheet_name, nrows=nrows, usecols=usecols)


def read_excel_upload(uploaded_file, sheet_name=0, nrows=None, usecols=None) -> pd.DataFrame:
    return _read_excel_bytes(uploaded_file.getvalue(), uploaded_file.name, sheet_name=sheet_name, nrows=nrows, usecols=usecols)


def read_sap_text_bytes(data: bytes) -> pd.DataFrame:
    # SAP list export usually UTF-16 LE, tab separated, with header line after a few title rows.
    decoded = None
    for enc in ["utf-16", "latin1", "utf-8-sig", "utf-8"]:
        try:
            decoded = data.decode(enc, errors="ignore")
            if "Nº pers." in decoded or "N° pers." in decoded or "CC-n." in decoded:
                break
        except Exception:
            continue
    if decoded is None:
        decoded = data.decode("latin1", errors="ignore")
    lines = decoded.splitlines(True)
    header_idx = None
    for i, line in enumerate(lines):
        if ("Nº pers." in line or "N° pers." in line or "N pers" in line) and ("CC-n." in line or "Per.para" in line):
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("No encontré encabezado SAP en archivo texto")
    text = "".join(lines[header_idx:])
    df = pd.read_csv(io.StringIO(text), sep="\t", dtype=str, engine="python")
    df = df.dropna(axis=1, how="all").dropna(how="all")
    # remover columnas unnamed/vacías
    keep = [c for c in df.columns if not str(c).startswith("Unnamed") and str(c).strip() != ""]
    df = df[keep].copy()
    return df


def read_any_upload(uploaded_file, sheet_name=0) -> pd.DataFrame:
    data = uploaded_file.getvalue()
    name = uploaded_file.name
    ext = os.path.splitext(name)[1].lower()
    if ext in [".txt", ".csv"] or ext == ".xls":
        # Muchas salidas SAP vienen .XLS pero realmente son texto UTF-16 tabulado.
        try:
            return read_sap_text_bytes(data)
        except Exception:
            pass
    return read_excel_upload(uploaded_file, sheet_name=sheet_name)


def list_sheets(uploaded_file) -> List[str]:
    data = uploaded_file.getvalue(); name = uploaded_file.name; ext = os.path.splitext(name)[1].lower()
    if ext == ".xlsb":
        xl = pd.ExcelFile(io.BytesIO(data), engine="pyxlsb")
    else:
        xl = pd.ExcelFile(io.BytesIO(data))
    return xl.sheet_names

# ==============================
# Homologación
# ==============================
def read_detalle_horas(detalle_file) -> Tuple[Dict[str, str], Dict[str, str], pd.DataFrame]:
    """Retorna (concepto->tipo_hora, funcion_key->cargo_homologado, tabla)."""
    concept_map = dict(CONCEPTOS)
    func_map: Dict[str, str] = {}
    audit_rows = []
    if detalle_file is not None:
        try:
            df = read_excel_upload(detalle_file, sheet_name="Homologación", nrows=5000)
        except Exception:
            df = read_excel_upload(detalle_file, sheet_name=0, nrows=5000)
        df = drop_duplicated_columns(df).dropna(how="all")
        # limpiar filas vacías masivas de xlsb
        df = df.loc[df.notna().sum(axis=1) > 0].copy()
        c_cc = find_col(df, ["CC-n.", "CC-n", "Concepto"], False)
        c_hora = find_col(df, ["Hora", "Tipo hora"], False)
        if c_cc and c_hora:
            tmp = df[[c_cc, c_hora]].dropna(how="all").copy()
            for _, r in tmp.iterrows():
                cc = clean_concept(r.get(c_cc))
                hora = clean_text(r.get(c_hora))
                if cc and hora:
                    concept_map[cc] = hora
        # funciones a cargos homologados
        c_func_code = find_col(df, ["Función", "Funcion"], False)
        c_func_text = find_col(df, ["Función_4", "Funcion_4", "Nombre función", "Nombre funcion"], False)
        c_cargo = find_col(df, ["Cargo"], False)
        if c_cargo:
            for _, r in df.iterrows():
                cargo = clean_text(r.get(c_cargo))
                if not cargo:
                    continue
                if c_func_code:
                    fc = clean_code(r.get(c_func_code))
                    if fc:
                        func_map[norm_key(fc)] = cargo
                        audit_rows.append({"origen": "Detalle Horas", "tipo_llave": "codigo_funcion", "llave": fc, "funcion": clean_text(r.get(c_func_text)) if c_func_text else "", "cargo_homologado": cargo})
                if c_func_text:
                    ft = clean_text(r.get(c_func_text))
                    if ft:
                        func_map[norm_key(ft)] = cargo
                        audit_rows.append({"origen": "Detalle Horas", "tipo_llave": "texto_funcion", "llave": ft, "funcion": ft, "cargo_homologado": cargo})
    # manual conceptos
    concept_map.setdefault("Y350", "Compensatorio")
    audit = pd.DataFrame(audit_rows).drop_duplicates() if audit_rows else pd.DataFrame(columns=["origen","tipo_llave","llave","funcion","cargo_homologado"])
    return concept_map, func_map, audit


def fuzzy_cargo_from_function(funcion: Any, area_negocio: str = "") -> str:
    s = norm_key(funcion)
    a = norm_key(area_negocio)
    if not s:
        return "Sin homologar"
    if "APRENDIZ" in s:
        return "Aprendiz"
    if "PART TIME" in s or "PARTTIME" in s:
        return "Part time"
    if "JEFE" in s and "TIENDA" in s:
        return "Jefe Tienda"
    if "SUPERVISOR" in s and ("JR" in s or "JUNIOR" in s):
        return "Supervisor Jr"
    if "SUPERVISOR" in s and "TIENDA" in s:
        return "Supervisor Tienda"
    if "OPERADOR" in s and "TIENDA" in s:
        return "Operador Tienda"
    if "MONTACARGA" in s and ("BDC" in s or a == "BDC"):
        return "Op . Montacarga BDC"
    if "MONTACARGA" in s:
        return "Montacarga Cedi"
    if "CENTRO DE DISTRIB" in s or "CEDI" in s or "DISTRIBUCION" in s or ("OPERADOR" in s and a == "CEDI"):
        return "Op. Cedi"
    if a == "OFICINA SOPORTE" or "ANALISTA" in s or "COORDINADOR" in s or "ESPECIALISTA" in s or "ASISTENTE" in s:
        return "Oficina Soporte"
    return "Sin homologar"


def homologate_function(func_code: Any, func_text: Any, func_map: Dict[str, str], area_negocio: str = "") -> Tuple[str, str]:
    fc = clean_code(func_code)
    ft = clean_text(func_text)
    if fc and norm_key(fc) in func_map:
        return func_map[norm_key(fc)], "funcion_codigo_detalle"
    if ft and norm_key(ft) in func_map:
        return func_map[norm_key(ft)], "funcion_texto_detalle"
    # si el texto que llega ya es un cargo homologado conocido
    if ft:
        fuzzy = fuzzy_cargo_from_function(ft, area_negocio)
        if fuzzy != "Sin homologar":
            return fuzzy, "funcion_fuzzy"
        return ft, "funcion_sin_mapeo"
    return "Sin homologar", "sin_funcion"


def build_position_function_master(hc_df: pd.DataFrame, poshom_df: Optional[pd.DataFrame] = None) -> Tuple[Dict[str, Dict[str, str]], pd.DataFrame]:
    rows = []
    def add(pos_code, pos_text, func_code, func_text, source, periodo=""):
        pos_code = clean_code(pos_code); pos_text = clean_text(pos_text); func_code = clean_code(func_code); func_text = clean_text(func_text)
        if not func_text and not func_code:
            return
        for key_val, key_type in [(pos_text, "posicion_texto"), (pos_code, "posicion_codigo"), (func_text, "funcion_texto"), (func_code, "funcion_codigo")]:
            if key_val:
                rows.append({
                    "key": norm_key(key_val), "llave_original": key_val, "tipo_llave": key_type,
                    "posicion_codigo": pos_code, "posicion_nombre": pos_text,
                    "funcion_codigo": func_code, "funcion_nombre": func_text,
                    "origen": source, "periodo": periodo,
                })
    if poshom_df is not None and not poshom_df.empty:
        poshom_df = drop_duplicated_columns(poshom_df)
        cp_code = find_col(poshom_df, ["Posición", "Posicion"], False)
        cp_text = find_col(poshom_df, ["Posición_3", "Posicion_3", "Nombre posición", "Nombre posicion", "Posición nombre"], False)
        cf_code = find_col(poshom_df, ["Función", "Funcion"], False)
        cf_text = find_col(poshom_df, ["Función_4", "Funcion_4", "Nombre función", "Nombre funcion"], False)
        for _, r in poshom_df.iterrows():
            add(r.get(cp_code) if cp_code else "", r.get(cp_text) if cp_text else "", r.get(cf_code) if cf_code else "", r.get(cf_text) if cf_text else "", "Posiciones_homologadas")
    if hc_df is not None and not hc_df.empty:
        for _, r in hc_df.iterrows():
            add(r.get("posicion_codigo", ""), r.get("posicion_nombre", ""), r.get("funcion_codigo", ""), r.get("funcion_nombre", ""), "Headcount", r.get("periodo_novedad", ""))
    if not rows:
        return {}, pd.DataFrame()
    raw = pd.DataFrame(rows)
    # Prioridad: Posiciones_homologadas > Headcount; dentro de llave, tomar la combinación más frecuente
    raw["prioridad"] = np.where(raw["origen"].eq("Posiciones_homologadas"), 0, 1)
    raw["combo"] = raw["funcion_codigo"].fillna("").astype(str) + "|" + raw["funcion_nombre"].fillna("").astype(str)
    counts = raw.groupby(["key", "combo"], dropna=False).size().reset_index(name="conteo")
    raw = raw.merge(counts, on=["key", "combo"], how="left")
    raw = raw.sort_values(["key", "prioridad", "conteo"], ascending=[True, True, False])
    best = raw.drop_duplicates("key", keep="first").copy()
    mapping = {
        r["key"]: {
            "funcion_codigo": r.get("funcion_codigo", ""),
            "funcion_nombre": r.get("funcion_nombre", ""),
            "posicion_codigo": r.get("posicion_codigo", ""),
            "posicion_nombre": r.get("posicion_nombre", ""),
            "origen_maestro": r.get("origen", ""),
        }
        for _, r in best.iterrows()
    }
    return mapping, raw.drop(columns=["combo"], errors="ignore")


def resolve_function_from_master(value: Any, master: Dict[str, Dict[str, str]]) -> Tuple[str, str, str]:
    key = norm_key(value)
    if key and key in master:
        info = master[key]
        return info.get("funcion_codigo", ""), info.get("funcion_nombre", ""), info.get("origen_maestro", "maestro")
    return "", "", "sin_maestro"

# ==============================
# Procesadores de fuentes
# ==============================
def process_headcount(files: List[Any], concept_map: Dict[str, str], func_map: Dict[str, str]) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, List[Dict[str, Any]]]:
    all_rows = []
    alerts = []
    for f in files or []:
        try:
            df = read_any_upload(f)
            df = drop_duplicated_columns(df)
            c_sap = find_col(df, ["Nº pers.", "N° pers.", "SAP"], True)
            c_status = find_col(df, ["Status ocupación", "Status ocupacion", "Status"], False)
            c_div = find_col(df, ["División de personal", "Division de personal"], False)
            c_area_pers = find_col(df, ["Área de personal", "Area de personal"], False)
            c_area_nom = find_col(df, ["Área de nómina", "Area de nomina", "Texto área nómina"], False)
            c_ceco = find_col(df, ["Ce.coste", "Ce coste", "CECO"], False)
            c_pos_code = find_col(df, ["Posición", "Posicion"], False)
            c_pos_text = find_col(df, ["Posición.1", "Posicion.1", "Posición_3", "Posicion_3"], False)
            c_func_code = find_col(df, ["Función", "Funcion"], False)
            c_func_text = find_col(df, ["Función.1", "Funcion.1", "Función_4", "Funcion_4"], False)
            file_name = getattr(f, "name", "")
            periodo = normalize_period_value("", file_name)
            out = pd.DataFrame(index=df.index)
            out["archivo"] = file_name
            out["periodo_novedad"] = periodo
            out["sap"] = df[c_sap].apply(clean_sap)
            out["status"] = df[c_status].apply(clean_text) if c_status else ""
            out["division"] = df[c_div].apply(clean_text) if c_div else ""
            out["area_personal"] = df[c_area_pers].apply(clean_text) if c_area_pers else ""
            out["area_nomina"] = df[c_area_nom].apply(clean_text) if c_area_nom else ""
            out["ceco"] = df[c_ceco].apply(clean_code) if c_ceco else ""
            out["posicion_codigo"] = df[c_pos_code].apply(clean_code) if c_pos_code else ""
            out["posicion_nombre"] = df[c_pos_text].apply(clean_text) if c_pos_text else ""
            out["funcion_codigo"] = df[c_func_code].apply(clean_code) if c_func_code else ""
            out["funcion_nombre"] = df[c_func_text].apply(clean_text) if c_func_text else ""
            out["area_negocio"] = [classify_area(ceco, div, "", area, pos) for ceco, div, area, pos in zip(out["ceco"], out["division"], out["area_nomina"], out["posicion_nombre"])]
            out["manager_excluido"] = [is_manager_excl(ap, an, pos) for ap, an, pos in zip(out["area_personal"], out["area_nomina"], out["posicion_nombre"])]
            out["cargo_homologado"], out["metodo_cargo"] = zip(*[homologate_function(fc, ft, func_map, area) for fc, ft, area in zip(out["funcion_codigo"], out["funcion_nombre"], out["area_negocio"])])
            all_rows.append(out)
            alerts.append({"tipo":"Cargue", "mensaje":f"Headcount {f.name}: {len(out):,} registros leídos"})
        except Exception as e:
            alerts.append({"tipo":"Error", "mensaje":f"Error procesando Headcount {getattr(f,'name','archivo')}: {e}"})
    hc_full = pd.concat(all_rows, ignore_index=True) if all_rows else pd.DataFrame()
    if hc_full.empty:
        return hc_full, pd.DataFrame(), pd.DataFrame(), alerts
    if "periodo_novedad" in hc_full.columns:
        hc_full["periodo_novedad"] = hc_full.apply(lambda r: normalize_period_value(r.get("periodo_novedad", ""), r.get("archivo", "")), axis=1)
    missing_period = int(hc_full["periodo_novedad"].eq("").sum()) if "periodo_novedad" in hc_full.columns else len(hc_full)
    if missing_period:
        alerts.append({"tipo":"Headcount", "mensaje":f"Headcount: {missing_period:,} registros sin mes detectado. Revise nombres de archivo; se requiere formato 012026, 01.2026 o 202601."})
    hc_full = hc_full[hc_full["sap"].ne("")].copy()
    excl = hc_full[hc_full["manager_excluido"]].copy()
    hc_valid = hc_full[~hc_full["manager_excluido"]].copy()
    # HC por mes, área, cargo. Contar SAP único.
    hc_group = hc_valid.groupby(KEY_HC, dropna=False)["sap"].nunique().reset_index(name="hc")
    hc_group = ensure_key_types(hc_group, KEY_HC)
    if len(excl):
        alerts.append({"tipo":"Headcount", "mensaje":f"Headcount: se excluyeron {len(excl):,} registros Manager I-IV/no aplican horas"})
    return hc_full, hc_group, excl, alerts


def finalize_headcount_with_master(hc_full: pd.DataFrame, master: Dict[str, Dict[str, str]], func_map: Dict[str, str]) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, List[Dict[str, Any]]]:
    """Recalcula el HC usando la regla correcta:
    periodo + posición del Headcount -> función -> cargo homologado.
    El HC se cuenta después de homologar posiciones por periodo y excluyendo Manager I-IV.
    """
    alerts: List[Dict[str, Any]] = []
    if hc_full is None or hc_full.empty:
        return pd.DataFrame(), pd.DataFrame(columns=KEY_HC + ["hc"]), pd.DataFrame(), alerts

    hc = drop_duplicated_columns(hc_full.copy())
    for c in ["periodo_novedad", "sap", "posicion_codigo", "posicion_nombre", "funcion_codigo", "funcion_nombre", "area_negocio", "manager_excluido"]:
        if c not in hc.columns:
            hc[c] = ""

    hc["periodo_novedad"] = hc.apply(lambda r: normalize_period_value(r.get("periodo_novedad", ""), r.get("archivo", "")), axis=1)
    hc["sap"] = hc["sap"].apply(clean_sap)
    hc["posicion_codigo"] = hc["posicion_codigo"].apply(clean_code)
    hc["posicion_nombre"] = hc["posicion_nombre"].apply(clean_text)
    hc["funcion_codigo"] = hc["funcion_codigo"].apply(clean_code)
    hc["funcion_nombre"] = hc["funcion_nombre"].apply(clean_text)
    hc["area_negocio"] = hc["area_negocio"].replace("", "Sin clasificar").fillna("Sin clasificar")

    # Resolver función desde maestro de posiciones, con fallback a función propia del Headcount
    resolved = []
    for _, r in hc.iterrows():
        fc, fn, origin = resolve_function_from_master(r.get("posicion_nombre", ""), master)
        if not fn:
            fc, fn, origin = resolve_function_from_master(r.get("posicion_codigo", ""), master)
        if not fn:
            fc = clean_code(r.get("funcion_codigo", ""))
            fn = clean_text(r.get("funcion_nombre", ""))
            origin = "funcion_headcount" if fn or fc else "sin_funcion"
        cargo, metodo = homologate_function(fc, fn, func_map, r.get("area_negocio", ""))
        resolved.append((fc, fn, origin, cargo, metodo))

    if resolved:
        hc[["funcion_codigo_final", "funcion_nombre_final", "origen_funcion", "cargo_homologado", "metodo_cargo"]] = pd.DataFrame(resolved, index=hc.index)
    else:
        hc["funcion_codigo_final"] = ""
        hc["funcion_nombre_final"] = ""
        hc["origen_funcion"] = ""
        hc["cargo_homologado"] = "Sin homologar"
        hc["metodo_cargo"] = ""

    excl = hc[hc["manager_excluido"].astype(bool)].copy() if "manager_excluido" in hc.columns else pd.DataFrame()
    valid = hc[(~hc["manager_excluido"].astype(bool)) & hc["periodo_novedad"].ne("")].copy() if "manager_excluido" in hc.columns else hc[hc["periodo_novedad"].ne("")].copy()

    # HC = conteo de posiciones/personas homologadas por periodo. Preferir SAP único; si no hay SAP, contar filas.
    if "sap" in valid.columns and valid["sap"].ne("").any():
        hc_group = valid[valid["sap"].ne("")].groupby(KEY_HC, dropna=False)["sap"].nunique().reset_index(name="hc")
    else:
        hc_group = valid.groupby(KEY_HC, dropna=False).size().reset_index(name="hc")
    hc_group = ensure_key_types(hc_group, KEY_HC)
    hc_group["hc"] = hc_group["hc"].fillna(0).astype(float)

    missing = int(hc["periodo_novedad"].eq("").sum())
    if missing:
        alerts.append({"tipo":"Headcount", "mensaje":f"Headcount: {missing:,} registros sin periodo; no entran al conteo HC."})
    if len(excl):
        alerts.append({"tipo":"Headcount", "mensaje":f"Headcount: se excluyeron {len(excl):,} registros Manager I-IV/no aplican horas"})
    sin_cargo = int(valid["cargo_homologado"].isin(["", "Sin homologar", "Sin cargo"]).sum())
    if sin_cargo:
        alerts.append({"tipo":"Homologación", "mensaje":f"Headcount: {sin_cargo:,} registros sin cargo homologado después de Posición → Función."})
    alerts.append({"tipo":"Headcount", "mensaje":f"HC calculado desde posiciones homologadas por periodo: {len(hc_group):,} combinaciones Mes+Área+Cargo."})
    return hc, hc_group, excl, alerts


def add_function_and_cargo(df: pd.DataFrame, source: str, master: Dict[str, Dict[str, str]], func_map: Dict[str, str], hc_full: pd.DataFrame, alerts: List[Dict[str, Any]], prefer_sap: bool = False) -> pd.DataFrame:
    """Añade función_periodo y cargo_homologado. Prioridad:
    1) SAP+periodo en HC si prefer_sap=True y está disponible.
    2) función directa si viene.
    3) cargo/posición contra Maestro Posición→Función.
    4) fallback fuzzy.
    """
    df = df.copy()
    for c in ["sap", "periodo_novedad", "posicion_original", "funcion_codigo", "funcion_nombre", "area_negocio"]:
        if c not in df.columns:
            df[c] = ""
    df["funcion_codigo_final"] = df["funcion_codigo"].apply(clean_code)
    df["funcion_nombre_final"] = df["funcion_nombre"].apply(clean_text)
    df["origen_funcion"] = np.where(df["funcion_nombre_final"].ne("") | df["funcion_codigo_final"].ne(""), "funcion_fuente", "")

    # SAP + periodo en HC, útil sobre todo para provisión
    if prefer_sap and hc_full is not None and not hc_full.empty:
        hc_sap = hc_full[["sap", "periodo_novedad", "funcion_codigo", "funcion_nombre", "posicion_nombre", "area_negocio"]].drop_duplicates(["sap","periodo_novedad"])
        before_cols = list(df.columns)
        df = df.merge(hc_sap, on=["sap","periodo_novedad"], how="left", suffixes=("", "_hc"))
        mask = (df["funcion_nombre_final"].eq("") | df["funcion_nombre_final"].isna()) & df["funcion_nombre_hc"].fillna("").ne("")
        df.loc[mask, "funcion_nombre_final"] = df.loc[mask, "funcion_nombre_hc"].apply(clean_text)
        df.loc[mask, "funcion_codigo_final"] = df.loc[mask, "funcion_codigo_hc"].apply(clean_code)
        df.loc[mask, "origen_funcion"] = "sap_periodo_headcount"
        # Completar área desde HC si venía vacía/sin clasificar
        mask_area = df["area_negocio"].isin(["", "Sin clasificar"]) & df.get("area_negocio_hc", pd.Series([""]*len(df))).fillna("").ne("")
        if "area_negocio_hc" in df.columns:
            df.loc[mask_area, "area_negocio"] = df.loc[mask_area, "area_negocio_hc"]
        # limpiar columnas _hc auxiliares excepto información que no moleste
        drop_cols = [c for c in df.columns if c.endswith("_hc")]
        df = df.drop(columns=drop_cols, errors="ignore")

    # Usar maestro posición -> función si aún no hay función
    for idx in df.index[df["funcion_nombre_final"].fillna("").eq("") & df["posicion_original"].fillna("").ne("")]:
        fc, ft, origin = resolve_function_from_master(df.at[idx, "posicion_original"], master)
        if ft or fc:
            df.at[idx, "funcion_codigo_final"] = fc
            df.at[idx, "funcion_nombre_final"] = ft
            df.at[idx, "origen_funcion"] = f"maestro_posicion_funcion:{origin}"

    # Si lo que viene como posicion_original es realmente función y existe en detalle, también sirve.
    for idx in df.index[df["funcion_nombre_final"].fillna("").eq("") & df["posicion_original"].fillna("").ne("")]:
        txt = df.at[idx, "posicion_original"]
        if norm_key(txt) in func_map:
            df.at[idx, "funcion_nombre_final"] = txt
            df.at[idx, "origen_funcion"] = "texto_ya_es_funcion"

    # Homologar función -> cargo homologado
    cargos = []
    methods = []
    for fc, ft, area in zip(df["funcion_codigo_final"], df["funcion_nombre_final"], df["area_negocio"]):
        cargo, metodo = homologate_function(fc, ft, func_map, area)
        cargos.append(cargo)
        methods.append(metodo)
    df["cargo_homologado"] = cargos
    df["metodo_cargo"] = methods

    pendientes = df[df["cargo_homologado"].isin(["", "Sin homologar"])]
    if len(pendientes):
        alerts.append({"tipo":"Homologación", "mensaje":f"{source}: {len(pendientes):,} registros quedaron sin cargo homologado. Revise hoja Pendientes_Homologacion."})
    return df


def process_pagado(cc_files: List[Any], comp_files: List[Any], concept_map: Dict[str, str], func_map: Dict[str, str], master: Dict[str, Dict[str, str]], hc_full: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, List[Dict[str, Any]]]:
    rows = []
    alerts = []
    all_files = [(f, "CCNómina") for f in (cc_files or [])] + [(f, "Compensatorios") for f in (comp_files or [])]
    for f, source in all_files:
        try:
            df = read_any_upload(f)
            df = drop_duplicated_columns(df)
            c_sap = find_col(df, ["Nº pers.", "N° pers.", "SAP"], False)
            c_period = find_col(df, ["Per.para", "Periodo", "Periodo para nómina", "Periodo para nomina"], False)
            c_fecha_pago = find_col(df, ["Fecha pago", "Fecha de pago"], False)
            c_concept = find_col(df, ["CC-n.", "CC-n", "Concepto", "Valores"], True)
            c_text = find_col(df, ["Texto expl.CC-nómina", "Texto expl.CC-nomina", "Txt.expl.", "Texto concepto"], False)
            c_qty = find_col(df, ["Cantidad", "Total"], True)
            c_value = find_col(df, ["Importe", "     Importe", "   Importe", "Valor"], True)
            c_ceco = find_col(df, ["Ce.coste", "CECO", "Ce coste"], False)
            c_div = find_col(df, ["Texto división de personal", "Texto division de personal", "División de personal"], False)
            c_area_nom = find_col(df, ["Texto área nómina", "Texto area nomina", "Área de nómina", "Area de nomina"], False)
            c_func_code = find_col(df, ["Función", " Funcion", "Función ", "Funcion"], False)
            c_func_text = find_col(df, ["Denominación función", "Denominacion funcion", "Función.1", "Funcion.1"], False)

            out = pd.DataFrame(index=df.index)
            out["fuente"] = source
            out["archivo"] = f.name
            out["sap"] = df[c_sap].apply(clean_sap) if c_sap else ""
            out["periodo_pago"] = df[c_period].apply(lambda x: parse_period_any(x, f.name)) if c_period else parse_period_any("", f.name)
            # si no detecta periodo_pago por columna, intenta fecha pago o nombre
            if c_fecha_pago:
                mask_empty = out["periodo_pago"].eq("")
                out.loc[mask_empty, "periodo_pago"] = df.loc[mask_empty, c_fecha_pago].apply(lambda x: parse_period_any(x, f.name))
            out["periodo_novedad"] = out["periodo_pago"].apply(prev_period)
            out["concepto"] = df[c_concept].apply(clean_concept)
            out["concepto"] = out["concepto"].map(lambda x: INTERFAZ_MAP.get(x, x))
            out["tipo_hora"] = out["concepto"].map(concept_map).fillna(out["concepto"].map(CONCEPTOS)).fillna("Sin tipo hora")
            out["cantidad_pagada"] = df[c_qty].apply(parse_number)
            out["valor_pagado"] = df[c_value].apply(parse_number)
            out["ceco"] = df[c_ceco].apply(clean_code) if c_ceco else ""
            div = df[c_div].apply(clean_text) if c_div else pd.Series([""]*len(df))
            area_nom = df[c_area_nom].apply(clean_text) if c_area_nom else pd.Series([""]*len(df))
            func_text = df[c_func_text].apply(clean_text) if c_func_text else pd.Series([""]*len(df))
            out["funcion_codigo"] = df[c_func_code].apply(clean_code) if c_func_code else ""
            out["funcion_nombre"] = func_text
            out["posicion_original"] = func_text  # En CC nómina normalmente viene función; si no, se tratará como texto
            out["area_negocio"] = [classify_area(ceco, d, "", an, ft) for ceco, d, an, ft in zip(out["ceco"], div, area_nom, func_text)]
            out = out[out["concepto"].isin(CONCEPTOS_SET)].copy()
            out = out[(out["cantidad_pagada"].abs() > 0) | (out["valor_pagado"].abs() > 0)].copy()
            out = add_function_and_cargo(out, source, master, func_map, hc_full, alerts, prefer_sap=False)
            rows.append(out)
            alerts.append({"tipo":"Cargue", "mensaje":f"{source} {f.name}: {len(out):,} registros útiles procesados"})
        except Exception as e:
            alerts.append({"tipo":"Error", "mensaje":f"Error procesando {source} {getattr(f,'name','archivo')}: {e}"})
    full = pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()
    if full.empty:
        return full, pd.DataFrame(), alerts
    agg = full.groupby(KEY_DETAIL, dropna=False).agg(cantidad_pagada=("cantidad_pagada","sum"), valor_pagado=("valor_pagado","sum")).reset_index()
    agg = ensure_key_types(agg, KEY_DETAIL)
    return full, agg, alerts



def _find_idx_from_headers(headers: List[str], candidates: List[str], required: bool = False) -> Optional[int]:
    norm_headers = {norm_key(h): i for i, h in enumerate(headers)}
    cand_norm = [norm_key(c) for c in candidates]
    for cn in cand_norm:
        if cn in norm_headers:
            return norm_headers[cn]
    for cn in cand_norm:
        matches = [(i, h) for h, i in norm_headers.items() if cn and (cn in h or h in cn)]
        if matches:
            return sorted(matches, key=lambda x: len(x[1]))[0][0]
    if required:
        raise ValueError(f"No encontré columna para: {candidates}")
    return None


def _row_get(row: Tuple[Any, ...], idx: Optional[int], default: Any = "") -> Any:
    if idx is None:
        return default
    if idx >= len(row):
        return default
    v = row[idx]
    return default if v is None else v


def _provision_from_dataframe_fast(df: pd.DataFrame, file_name: str, concept_map: Dict[str, str]) -> Tuple[pd.DataFrame, int]:
    """Procesa provisión desde CSV/Parquet/DataFrame de forma vectorizada y agrega antes de homologar."""
    df = drop_duplicated_columns(df)
    c_source = find_col(df, ["Source.Name", "Source Name", "MES", "Periodo"], False)
    c_ceco = find_col(df, ["CECO", "Ce.coste", "Ce coste"], False)
    c_tipo = find_col(df, ["TIPO", "Tipo"], False)
    c_cargo = find_col(df, ["CARGO", "Cargo"], True)
    c_concept = find_col(df, ["Valores", "Concepto", "CC-n."], True)
    c_qty = find_col(df, ["Total", "Cantidad"], True)
    c_value = find_col(df, ["PROVISIÓN", "PROVISION", "Provisión", "Provision"], True)
    c_region = find_col(df, ["REGION", "Región", "Division", "División"], False)

    base = pd.DataFrame()
    base["fuente"] = "Provisión"
    base["archivo"] = file_name
    base["periodo_novedad"] = df[c_source].apply(lambda x: parse_period_any(x, file_name)) if c_source else parse_period_any("", file_name)
    base["ceco"] = df[c_ceco].apply(clean_code) if c_ceco else ""
    cargo = df[c_cargo].apply(clean_text)
    tipo = df[c_tipo].apply(clean_text) if c_tipo else pd.Series([""] * len(df), index=df.index)
    region = df[c_region].apply(clean_text) if c_region else pd.Series([""] * len(df), index=df.index)
    base["posicion_original"] = cargo
    base["concepto"] = df[c_concept].apply(clean_concept)
    base["tipo_hora"] = base["concepto"].map(concept_map).fillna(base["concepto"].map(CONCEPTOS)).fillna("Sin tipo hora")
    base["cantidad_provisionada"] = df[c_qty].apply(parse_number)
    base["valor_provisionado"] = df[c_value].apply(parse_number)
    base["area_negocio"] = [classify_area(ceco, reg, ti, "", cg) for ceco, reg, ti, cg in zip(base["ceco"], region, tipo, cargo)]
    base = base[base["concepto"].isin(CONCEPTOS_SET)].copy()
    base = base[(base["cantidad_provisionada"].abs() > 0) | (base["valor_provisionado"].abs() > 0)].copy()
    if base.empty:
        return base, len(df)
    group_cols = ["fuente", "archivo", "periodo_novedad", "area_negocio", "ceco", "posicion_original", "concepto", "tipo_hora"]
    pre = base.groupby(group_cols, dropna=False).agg(
        cantidad_provisionada=("cantidad_provisionada", "sum"),
        valor_provisionado=("valor_provisionado", "sum"),
    ).reset_index()
    return pre, len(base)


def _read_provision_streaming_excel(file_obj, concept_map: Dict[str, str]) -> Tuple[pd.DataFrame, int]:
    """Lee provisión .xlsx/.xlsm en modo streaming con openpyxl.

    Evita cargar toda la hoja en memoria como pandas.read_excel. Agrega mientras lee.
    Esto es clave para Streamlit Cloud cuando el consolidado de provisión es grande.
    """
    from openpyxl import load_workbook

    data = file_obj.getvalue()
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    # Buscar hoja Horas_Provisión de forma flexible
    sheet_name = None
    for sn in wb.sheetnames:
        nk = norm_key(sn)
        if "HORAS" in nk and "PROVISION" in nk:
            sheet_name = sn
            break
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    header = None
    header_row_number = None
    # Buscar encabezado en primeras 30 filas
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        vals = [clean_text(v) for v in row]
        nk_join = " ".join(norm_key(v) for v in vals if v)
        if ("CARGO" in nk_join and ("VALORES" in nk_join or "CONCEPTO" in nk_join) and ("PROVISION" in nk_join or "TOTAL" in nk_join)):
            header = vals
            header_row_number = i
            break
        if i >= 30:
            break
    if header is None:
        wb.close()
        raise ValueError("No encontré encabezado válido en la provisión. Debe tener CARGO, Valores/Concepto y PROVISIÓN/Total.")

    i_source = _find_idx_from_headers(header, ["Source.Name", "Source Name", "MES", "Periodo"], False)
    i_ceco = _find_idx_from_headers(header, ["CECO", "Ce.coste", "Ce coste"], False)
    i_tipo = _find_idx_from_headers(header, ["TIPO", "Tipo"], False)
    i_cargo = _find_idx_from_headers(header, ["CARGO", "Cargo"], True)
    i_concept = _find_idx_from_headers(header, ["Valores", "Concepto", "CC-n."], True)
    i_qty = _find_idx_from_headers(header, ["Total", "Cantidad"], True)
    i_value = _find_idx_from_headers(header, ["PROVISIÓN", "PROVISION", "Provisión", "Provision"], True)
    i_region = _find_idx_from_headers(header, ["REGION", "Región", "Division", "División"], False)

    agg: Dict[Tuple[str, str, str, str, str, str, str, str], List[float]] = {}
    useful = 0
    scanned = 0
    for r_idx, row in enumerate(ws.iter_rows(min_row=header_row_number + 1, values_only=True), start=header_row_number + 1):
        scanned += 1
        concepto = clean_concept(_row_get(row, i_concept))
        if concepto not in CONCEPTOS_SET:
            continue
        qty = parse_number(_row_get(row, i_qty))
        val = parse_number(_row_get(row, i_value))
        if abs(qty) == 0 and abs(val) == 0:
            continue
        source_val = _row_get(row, i_source) if i_source is not None else ""
        periodo = parse_period_any(source_val, file_obj.name)
        ceco = clean_code(_row_get(row, i_ceco))
        tipo = clean_text(_row_get(row, i_tipo))
        cargo = clean_text(_row_get(row, i_cargo))
        region = clean_text(_row_get(row, i_region))
        area = classify_area(ceco, region, tipo, "", cargo)
        tipo_hora = concept_map.get(concepto) or CONCEPTOS.get(concepto, "Sin tipo hora")
        key = ("Provisión", file_obj.name, periodo, area, ceco, cargo, concepto, tipo_hora)
        if key not in agg:
            agg[key] = [0.0, 0.0]
        agg[key][0] += qty
        agg[key][1] += val
        useful += 1
    wb.close()
    if not agg:
        return pd.DataFrame(), useful
    records = []
    for (fuente, archivo, periodo, area, ceco, cargo, concepto, tipo_hora), (qty, val) in agg.items():
        records.append({
            "fuente": fuente, "archivo": archivo, "periodo_novedad": periodo,
            "area_negocio": area, "ceco": ceco, "posicion_original": cargo,
            "concepto": concepto, "tipo_hora": tipo_hora,
            "cantidad_provisionada": qty, "valor_provisionado": val,
        })
    return pd.DataFrame.from_records(records), useful


def _read_provision_preaggregated(file_obj, concept_map: Dict[str, str]) -> Tuple[pd.DataFrame, int, str]:
    """Devuelve provisión ya agregada antes de homologar. Soporta XLSX streaming, CSV y Parquet."""
    name = getattr(file_obj, "name", "archivo")
    ext = os.path.splitext(name)[1].lower()
    if ext == ".parquet":
        df = pd.read_parquet(io.BytesIO(file_obj.getvalue()))
        pre, useful = _provision_from_dataframe_fast(df, name, concept_map)
        return pre, useful, "parquet"
    if ext in [".csv", ".txt"]:
        # separador flexible: intenta ; y luego ,
        data = file_obj.getvalue()
        for sep in [";", ",", "\t"]:
            try:
                df = pd.read_csv(io.BytesIO(data), sep=sep, dtype=str, encoding="utf-8-sig")
                if df.shape[1] > 1:
                    pre, useful = _provision_from_dataframe_fast(df, name, concept_map)
                    return pre, useful, f"csv sep={sep!r}"
            except Exception:
                continue
        df = pd.read_csv(io.BytesIO(data), dtype=str, encoding="latin1")
        pre, useful = _provision_from_dataframe_fast(df, name, concept_map)
        return pre, useful, "csv"
    # Excel: streaming
    pre, useful = _read_provision_streaming_excel(file_obj, concept_map)
    return pre, useful, "excel_streaming"


def process_provision(files: List[Any], concept_map: Dict[str, str], func_map: Dict[str, str], master: Dict[str, Dict[str, str]], hc_full: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, List[Dict[str, Any]]]:
    """Procesa provisión V14 en modo streaming/agregado.

    - Para Excel usa openpyxl read_only y agrega fila a fila sin cargar toda la hoja.
    - Para CSV/Parquet usa lectura rápida.
    - Luego homologa únicamente las combinaciones agregadas.
    """
    rows = []
    alerts = []

    for f in files or []:
        try:
            pre, useful_rows, mode = _read_provision_preaggregated(f, concept_map)
            if pre.empty:
                alerts.append({"tipo": "Cargue", "mensaje": f"Provisión {f.name}: sin registros útiles después de filtrar conceptos/valores. Modo: {mode}."})
                continue

            func_codes = []
            func_names = []
            origins = []
            cargos = []
            methods = []
            for pos, area in zip(pre["posicion_original"], pre["area_negocio"]):
                fc, ft, origin = resolve_function_from_master(pos, master)
                if not ft and not fc and norm_key(pos) in func_map:
                    ft = clean_text(pos)
                    origin = "texto_ya_es_funcion"
                cargo_hom, metodo = homologate_function(fc, ft if ft else pos, func_map, area)
                func_codes.append(fc)
                func_names.append(ft if ft else clean_text(pos))
                origins.append(origin)
                cargos.append(cargo_hom)
                methods.append(metodo)
            pre["sap"] = ""
            pre["funcion_codigo_final"] = func_codes
            pre["funcion_nombre_final"] = func_names
            pre["origen_funcion"] = origins
            pre["cargo_homologado"] = cargos
            pre["metodo_cargo"] = methods

            pend = pre[pre["cargo_homologado"].isin(["", "Sin homologar", "Sin cargo"])]
            if len(pend):
                alerts.append({"tipo": "Homologación", "mensaje": f"Provisión {f.name}: {len(pend):,} combinaciones agregadas quedaron sin cargo homologado."})

            rows.append(pre)
            alerts.append({"tipo": "Cargue", "mensaje": f"Provisión {f.name}: {useful_rows:,} filas útiles leídas; {len(pre):,} combinaciones agregadas. Modo {mode}."})
        except Exception as e:
            alerts.append({"tipo": "Error", "mensaje": f"Error procesando provisión {getattr(f,'name','archivo')}: {e}"})

    full = pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()
    if full.empty:
        return full, pd.DataFrame(), alerts
    full = ensure_key_types(full, KEY_DETAIL)
    agg = full.groupby(KEY_DETAIL, dropna=False).agg(
        cantidad_provisionada=("cantidad_provisionada", "sum"),
        valor_provisionado=("valor_provisionado", "sum"),
    ).reset_index()
    agg = ensure_key_types(agg, KEY_DETAIL)
    return full, agg, alerts

def process_proyeccion(files: List[Any], concept_map: Dict[str, str], func_map: Dict[str, str], master: Dict[str, Dict[str, str]], hc_full: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, List[Dict[str, Any]]]:
    rows = []
    alerts = []
    for f in files or []:
        try:
            try:
                df = read_excel_upload(f, sheet_name="Horas_Proyección")
            except Exception:
                df = read_any_upload(f)
            df = drop_duplicated_columns(df)
            c_mes = find_col(df, ["MES", "Source.Name", "Periodo"], False)
            c_sap = find_col(df, ["SAP", "Nº pers."], False)
            c_func = find_col(df, ["Funcion", "Función", "Cargo"], False)
            c_ceco = find_col(df, ["Ce.coste", "CECO"], False)
            c_tipo = find_col(df, ["Tipo", "TIPO"], False)
            c_area_nom = find_col(df, ["Área de nómina", "Area de nomina"], False)
            base_cols = [c for c in [c_mes, c_sap, c_func, c_ceco, c_tipo, c_area_nom] if c]
            tmp_base = df[base_cols].copy() if base_cols else pd.DataFrame(index=df.index)
            long_rows = []
            for concepto in CONCEPTOS_SET:
                q_col = None; v_col = None
                # Preferir nombre exacto
                for c in df.columns:
                    if norm_key(c) == norm_key(f"{concepto}_Q"):
                        q_col = c
                    if norm_key(c) == norm_key(f"{concepto}_$") or str(c).strip().upper() == f"{concepto}_$":
                        v_col = c
                if q_col is None and f"{concepto}_Q" in df.columns:
                    q_col = f"{concepto}_Q"
                if v_col is None and f"{concepto}_$" in df.columns:
                    v_col = f"{concepto}_$"
                if q_col is None and v_col is None:
                    continue
                part = tmp_base.copy()
                part["concepto"] = concepto
                part["cantidad_proyectada"] = df[q_col].apply(parse_number) if q_col else 0.0
                part["valor_proyectado"] = df[v_col].apply(parse_number) if v_col else 0.0
                part = part[(part["cantidad_proyectada"].abs() > 0) | (part["valor_proyectado"].abs() > 0)]
                long_rows.append(part)
            if not long_rows:
                alerts.append({"tipo":"Proyección", "mensaje":f"Proyección {f.name}: no se encontraron columnas *_Q / *_$ con movimiento"})
                continue
            long = pd.concat(long_rows, ignore_index=True)
            out = pd.DataFrame(index=long.index)
            out["fuente"] = "Proyección"
            out["archivo"] = f.name
            out["periodo_novedad"] = long[c_mes].apply(lambda x: parse_period_any(x, f.name)) if c_mes else parse_period_any("", f.name)
            out["sap"] = long[c_sap].apply(clean_sap) if c_sap else ""
            out["ceco"] = long[c_ceco].apply(clean_code) if c_ceco else ""
            func = long[c_func].apply(clean_text) if c_func else pd.Series([""]*len(long))
            tipo = long[c_tipo].apply(clean_text) if c_tipo else pd.Series([""]*len(long))
            area_nom = long[c_area_nom].apply(clean_text) if c_area_nom else pd.Series([""]*len(long))
            out["posicion_original"] = func
            out["funcion_codigo"] = ""
            out["funcion_nombre"] = func  # Proyección suele traer función textual
            out["concepto"] = long["concepto"].apply(clean_concept)
            out["tipo_hora"] = out["concepto"].map(concept_map).fillna(out["concepto"].map(CONCEPTOS)).fillna("Sin tipo hora")
            out["cantidad_proyectada"] = long["cantidad_proyectada"].astype(float)
            out["valor_proyectado"] = long["valor_proyectado"].astype(float)
            out["area_negocio"] = [classify_area(ceco, "", ti, an, fu) for ceco, ti, an, fu in zip(out["ceco"], tipo, area_nom, func)]
            invalid_sap = out["sap"].eq("").sum()
            if invalid_sap:
                alerts.append({"tipo":"Proyección", "mensaje":f"{invalid_sap:,} registros de proyección tenían SAP inválido/Error; se homologaron por función/cargo."})
            out = add_function_and_cargo(out, "Proyección", master, func_map, hc_full, alerts, prefer_sap=False)
            rows.append(out)
            alerts.append({"tipo":"Cargue", "mensaje":f"Proyección {f.name}: {len(out):,} registros útiles procesados"})
        except Exception as e:
            alerts.append({"tipo":"Error", "mensaje":f"Error procesando proyección {getattr(f,'name','archivo')}: {e}"})
    full = pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()
    if full.empty:
        return full, pd.DataFrame(), alerts
    agg = full.groupby(KEY_DETAIL, dropna=False).agg(cantidad_proyectada=("cantidad_proyectada","sum"), valor_proyectado=("valor_proyectado","sum")).reset_index()
    agg = ensure_key_types(agg, KEY_DETAIL)
    return full, agg, alerts

# ==============================
# Agregación y reportes
# ==============================
def empty_agg(cols: List[str]) -> pd.DataFrame:
    return pd.DataFrame(columns=KEY_DETAIL + cols)


def aggregate_sources(pagado: pd.DataFrame, provision: pd.DataFrame, proyeccion: pd.DataFrame, hc: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    p = pagado.copy() if pagado is not None and not pagado.empty else empty_agg(["cantidad_pagada", "valor_pagado"])
    pr = provision.copy() if provision is not None and not provision.empty else empty_agg(["cantidad_provisionada", "valor_provisionado"])
    py = proyeccion.copy() if proyeccion is not None and not proyeccion.empty else empty_agg(["cantidad_proyectada", "valor_proyectado"])
    for df in [p, pr, py]:
        for k in KEY_DETAIL:
            if k not in df.columns:
                df[k] = ""
            df[k] = df[k].fillna("").astype(str)
    comp = p.merge(pr, on=KEY_DETAIL, how="outer")
    comp = comp.merge(py, on=KEY_DETAIL, how="outer")
    for c in ["cantidad_pagada", "valor_pagado", "cantidad_provisionada", "valor_provisionado", "cantidad_proyectada", "valor_proyectado"]:
        if c not in comp.columns:
            comp[c] = 0.0
        comp[c] = comp[c].fillna(0.0).astype(float)
    # HC por cargo/área/mes
    hc_use = hc.copy() if hc is not None and not hc.empty else pd.DataFrame(columns=KEY_HC + ["hc"])
    hc_use = ensure_key_types(hc_use, KEY_HC)
    hc_use["hc"] = hc_use.get("hc", 0).fillna(0).astype(float)
    comp = comp.merge(hc_use, on=KEY_HC, how="left")
    comp["hc"] = comp["hc"].fillna(0.0)
    comp = add_diff_cols(comp)
    # Ejecutivo por mes + área + cargo + concepto + tipo hora (sin CECO)
    exec_df = comp.groupby(KEY_EXEC, dropna=False).agg(
        cantidad_pagada=("cantidad_pagada","sum"),
        valor_pagado=("valor_pagado","sum"),
        cantidad_provisionada=("cantidad_provisionada","sum"),
        valor_provisionado=("valor_provisionado","sum"),
        cantidad_proyectada=("cantidad_proyectada","sum"),
        valor_proyectado=("valor_proyectado","sum"),
    ).reset_index()
    exec_df = exec_df.merge(hc_use, on=KEY_HC, how="left")
    exec_df["hc"] = exec_df["hc"].fillna(0.0)
    exec_df = add_diff_cols(exec_df)
    exec_sin_ceros = exec_df[(exec_df["valor_pagado"].abs() + exec_df["valor_provisionado"].abs() + exec_df["valor_proyectado"].abs() + exec_df["cantidad_pagada"].abs() + exec_df["cantidad_provisionada"].abs() + exec_df["cantidad_proyectada"].abs()) > 0].copy()
    cargo_df = exec_df.groupby(KEY_HC, dropna=False).agg(
        cantidad_pagada=("cantidad_pagada","sum"),
        valor_pagado=("valor_pagado","sum"),
        cantidad_provisionada=("cantidad_provisionada","sum"),
        valor_provisionado=("valor_provisionado","sum"),
        cantidad_proyectada=("cantidad_proyectada","sum"),
        valor_proyectado=("valor_proyectado","sum"),
        hc=("hc", "max"),
    ).reset_index()
    cargo_df = add_diff_cols(cargo_df)
    indicadores = cargo_df.copy()
    indicadores["horas_pagadas_por_hc"] = np.where(indicadores["hc"] > 0, indicadores["cantidad_pagada"] / indicadores["hc"], 0)
    indicadores["horas_provisionadas_por_hc"] = np.where(indicadores["hc"] > 0, indicadores["cantidad_provisionada"] / indicadores["hc"], 0)
    indicadores["horas_proyectadas_por_hc"] = np.where(indicadores["hc"] > 0, indicadores["cantidad_proyectada"] / indicadores["hc"], 0)
    indicadores["valor_pagado_por_hc"] = np.where(indicadores["hc"] > 0, indicadores["valor_pagado"] / indicadores["hc"], 0)
    indicadores["valor_provisionado_por_hc"] = np.where(indicadores["hc"] > 0, indicadores["valor_provisionado"] / indicadores["hc"], 0)
    indicadores["valor_proyectado_por_hc"] = np.where(indicadores["hc"] > 0, indicadores["valor_proyectado"] / indicadores["hc"], 0)
    resumen_mes = exec_df.groupby("periodo_novedad", dropna=False).agg(
        cantidad_pagada=("cantidad_pagada","sum"),
        valor_pagado=("valor_pagado","sum"),
        cantidad_provisionada=("cantidad_provisionada","sum"),
        valor_provisionado=("valor_provisionado","sum"),
        cantidad_proyectada=("cantidad_proyectada","sum"),
        valor_proyectado=("valor_proyectado","sum"),
    ).reset_index()
    resumen_mes = add_diff_cols(resumen_mes)
    for df in [comp, exec_df, exec_sin_ceros, cargo_df, indicadores, resumen_mes]:
        if "periodo_novedad" in df.columns:
            df["periodo_orden"] = df["periodo_novedad"].apply(period_sort_key)
            df.sort_values(["periodo_orden"] + [c for c in ["area_negocio","cargo_homologado","concepto","tipo_hora"] if c in df.columns], inplace=True)
            df.drop(columns=["periodo_orden"], inplace=True)
    return {
        "Detalle_Comparativo": comp,
        "Resumen_Ejecutivo": exec_df,
        "Resumen_Ejecutivo_Sin_Ceros": exec_sin_ceros,
        "Resumen_Cargo_Homologado": cargo_df,
        "Indicadores_HC": indicadores,
        "Resumen_Mes": resumen_mes,
    }


def add_diff_cols(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for c in ["valor_pagado","valor_provisionado","valor_proyectado","cantidad_pagada","cantidad_provisionada","cantidad_proyectada"]:
        if c not in df.columns:
            df[c] = 0.0
    df["dif_pagado_vs_provision"] = df["valor_pagado"] - df["valor_provisionado"]
    df["dif_pagado_vs_proyeccion"] = df["valor_pagado"] - df["valor_proyectado"]
    df["dif_cant_pagada_vs_provision"] = df["cantidad_pagada"] - df["cantidad_provisionada"]
    df["dif_cant_pagada_vs_proyeccion"] = df["cantidad_pagada"] - df["cantidad_proyectada"]
    df["desv_pct_vs_provision"] = np.where(df["valor_provisionado"].abs() > 0, df["dif_pagado_vs_provision"] / df["valor_provisionado"].abs() * 100, np.where(df["valor_pagado"].abs() > 0, 100.0, 0.0))
    df["desv_pct_vs_proyeccion"] = np.where(df["valor_proyectado"].abs() > 0, df["dif_pagado_vs_proyeccion"] / df["valor_proyectado"].abs() * 100, np.where(df["valor_pagado"].abs() > 0, 100.0, 0.0))
    return df


def build_alerts(report: Dict[str, pd.DataFrame], alerts: List[Dict[str, Any]], umbral: float = 15.0) -> pd.DataFrame:
    out = list(alerts)
    detalle = report.get("Detalle_Comparativo", pd.DataFrame())
    if not detalle.empty:
        p_sin_pr = detalle[(detalle["valor_pagado"].abs() > 0) & (detalle["valor_provisionado"].abs() == 0)]
        p_sin_py = detalle[(detalle["valor_pagado"].abs() > 0) & (detalle["valor_proyectado"].abs() == 0)]
        pr_sin_p = detalle[(detalle["valor_provisionado"].abs() > 0) & (detalle["valor_pagado"].abs() == 0)]
        py_sin_p = detalle[(detalle["valor_proyectado"].abs() > 0) & (detalle["valor_pagado"].abs() == 0)]
        hc_cero = detalle[(detalle["hc"].fillna(0) == 0) & ((detalle["valor_pagado"].abs()+detalle["valor_provisionado"].abs()+detalle["valor_proyectado"].abs()) > 0)]
        desv = detalle[(detalle["valor_pagado"].abs() > 0) & ((detalle["desv_pct_vs_provision"].abs() > umbral) | (detalle["desv_pct_vs_proyeccion"].abs() > umbral))]
        out += [
            {"tipo":"Cruce", "mensaje":f"Pagado sin provisión: {len(p_sin_pr):,} combinaciones"},
            {"tipo":"Cruce", "mensaje":f"Pagado sin proyección: {len(p_sin_py):,} combinaciones"},
            {"tipo":"Cruce", "mensaje":f"Provisión sin pagado: {len(pr_sin_p):,} combinaciones"},
            {"tipo":"Cruce", "mensaje":f"Proyección sin pagado: {len(py_sin_p):,} combinaciones"},
            {"tipo":"Headcount", "mensaje":f"Combinaciones con movimiento y HC en cero: {len(hc_cero):,}"},
            {"tipo":"Desviación", "mensaje":f"Combinaciones que superan {umbral:.0f}% de desviación: {len(desv):,}"},
        ]
    return pd.DataFrame(out)

# ==============================
# Presentación y Excel
# ==============================
FRIENDLY = {
    "periodo_novedad": "Mes novedad",
    "periodo_pago": "Mes pago",
    "area_negocio": "Área negocio",
    "cargo_homologado": "Cargo homologado",
    "ceco": "CECO",
    "concepto": "Concepto",
    "tipo_hora": "Tipo hora",
    "cantidad_pagada": "Cantidad pagada",
    "valor_pagado": "Pagado",
    "cantidad_provisionada": "Cantidad provisión",
    "valor_provisionado": "Provisión",
    "cantidad_proyectada": "Cantidad proyección",
    "valor_proyectado": "Proyección",
    "dif_pagado_vs_provision": "Dif. pagado vs provisión",
    "dif_pagado_vs_proyeccion": "Dif. pagado vs proyección",
    "dif_cant_pagada_vs_provision": "Dif. cant. vs provisión",
    "dif_cant_pagada_vs_proyeccion": "Dif. cant. vs proyección",
    "desv_pct_vs_provision": "% desv. vs provisión",
    "desv_pct_vs_proyeccion": "% desv. vs proyección",
    "hc": "HC",
    "horas_pagadas_por_hc": "Horas pagadas por HC",
    "horas_provisionadas_por_hc": "Horas provisión por HC",
    "horas_proyectadas_por_hc": "Horas proyección por HC",
    "valor_pagado_por_hc": "Pagado por HC",
    "valor_provisionado_por_hc": "Provisión por HC",
    "valor_proyectado_por_hc": "Proyección por HC",
}


def pretty_df(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy()
    # ordenar columnas importantes
    preferred = [c for c in ["periodo_novedad","area_negocio","cargo_homologado","ceco","concepto","tipo_hora","hc","cantidad_pagada","valor_pagado","cantidad_provisionada","valor_provisionado","cantidad_proyectada","valor_proyectado","dif_pagado_vs_provision","dif_pagado_vs_proyeccion","desv_pct_vs_provision","desv_pct_vs_proyeccion"] if c in out.columns]
    rest = [c for c in out.columns if c not in preferred]
    out = out[preferred + rest]
    for c in out.columns:
        lc = c.lower()
        if any(token in lc for token in ["valor", "pagado", "provision", "proyectado", "dif_pagado", "costo"]):
            # no formatear cantidades que tengan pagada/proyectada en nombre
            if "cantidad" not in lc and "desv" not in lc and "pct" not in lc and "horas" not in lc:
                if pd.api.types.is_numeric_dtype(out[c]):
                    out[c] = out[c].apply(format_money)
        if "cantidad" in lc or "horas" in lc:
            if pd.api.types.is_numeric_dtype(out[c]):
                out[c] = out[c].apply(format_qty)
        if c == "hc" or lc == "hc":
            if pd.api.types.is_numeric_dtype(out[c]):
                out[c] = out[c].apply(format_int)
        if "pct" in lc or "desv_pct" in lc:
            if pd.api.types.is_numeric_dtype(out[c]):
                out[c] = out[c].apply(format_pct)
    out = out.rename(columns=FRIENDLY)
    return out


def display_df(df: pd.DataFrame, name: str = "tabla", max_rows: int = MAX_SCREEN_ROWS):
    """Muestra tablas sin tumbar la app.

    Protecciones V14.5:
    - Limita filas en pantalla.
    - Elimina columnas duplicadas.
    - Evita errores de Arrow/Streamlit por tipos mixtos.
    - Si st.dataframe falla, muestra una tabla HTML simple y mantiene la app viva.
    """
    try:
        if df is None or len(df) == 0:
            st.info("No hay datos para mostrar con los filtros seleccionados.")
            return
        safe = df.copy()
        safe = safe.loc[:, ~safe.columns.duplicated()].copy()
        safe = safe.replace([np.inf, -np.inf], np.nan)
        total = len(safe)
        show = safe.head(max_rows).copy()
        if total > max_rows:
            st.warning(f"Se muestran las primeras {max_rows:,} filas de {total:,}. Descarga el Excel para revisar el detalle completo.")
        height = min(650, max(240, 38 * (len(show) + 1)))
        try:
            shown = pretty_df(show)
            shown = shown.loc[:, ~shown.columns.duplicated()].copy()
            st.dataframe(shown, width="stretch", height=int(height))
        except Exception as e:
            st.warning(f"La tabla se mostró en modo seguro porque Streamlit no pudo renderizarla normalmente: {type(e).__name__}")
            fallback = pretty_df(show).astype(str)
            st.dataframe(fallback, width="stretch", height=int(height))
    except Exception as e:
        st.error("No fue posible mostrar esta tabla, pero la app continúa funcionando.")
        st.caption(f"Detalle técnico: {type(e).__name__}: {e}")


def filter_df(df: pd.DataFrame, filters: Dict[str, List[str]]) -> pd.DataFrame:
    """Filtrado defensivo: nunca debe tumbar la app."""
    try:
        if df is None or len(df) == 0:
            return pd.DataFrame()
        out = df.copy()
        out = out.loc[:, ~out.columns.duplicated()].copy()
        for col, vals in filters.items():
            if vals and col in out.columns:
                vals_set = {str(v) for v in vals}
                out = out[out[col].astype(str).isin(vals_set)]
        return out
    except Exception as e:
        st.error("No fue posible aplicar los filtros. Se devuelve la base sin filtrar para no tumbar la app.")
        st.caption(f"Detalle técnico: {type(e).__name__}: {e}")
        return df.copy() if df is not None else pd.DataFrame()


def prepare_report_base(df: pd.DataFrame) -> pd.DataFrame:
    """Normaliza la base para filtros y evita errores por columnas/tipos."""
    if df is None or len(df) == 0:
        return pd.DataFrame()
    out = df.copy()
    out = out.loc[:, ~out.columns.duplicated()].copy()
    for c in ["periodo_novedad", "area_negocio", "cargo_homologado", "concepto", "tipo_hora"]:
        if c not in out.columns:
            out[c] = ""
        out[c] = out[c].fillna("").astype(str)
    numeric_cols = [
        "cantidad_pagada", "valor_pagado", "cantidad_provisionada", "valor_provisionado",
        "cantidad_proyectada", "valor_proyectado", "hc", "dif_pagado_vs_provision",
        "dif_pagado_vs_proyeccion", "dif_cant_pagada_vs_provision", "dif_cant_pagada_vs_proyeccion",
        "desv_pct_vs_provision", "desv_pct_vs_proyeccion"
    ]
    for c in numeric_cols:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c].apply(parse_number), errors="coerce").fillna(0.0)
    out["periodo_novedad"] = out["periodo_novedad"].apply(normalize_period_value)
    return out


def safe_unique_options(df: pd.DataFrame, col: str, sort_period: bool = False) -> List[str]:
    if df is None or df.empty or col not in df.columns:
        return []
    vals = [v for v in df[col].fillna("").astype(str).unique().tolist() if v != ""]
    if sort_period:
        return sorted(vals, key=period_sort_key)
    return sorted(vals)


def make_excel(report: Dict[str, pd.DataFrame], extras: Dict[str, pd.DataFrame]) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        wb = writer.book
        fmt_header = wb.add_format({"bold": True, "bg_color": ORANGE, "font_color": "white", "border": 1})
        fmt_money = wb.add_format({"num_format": '$ #,##0', "border": 1})
        fmt_qty = wb.add_format({"num_format": '#,##0.00', "border": 1})
        fmt_int = wb.add_format({"num_format": '#,##0', "border": 1})
        fmt_pct = wb.add_format({"num_format": '0.00%', "border": 1})
        fmt_text = wb.add_format({"border": 1})
        for sheet, df in {**report, **extras}.items():
            if df is None or df.empty:
                df = pd.DataFrame({"Mensaje": ["Sin datos"]})
            sheet_name = re.sub(r"[\[\]\:\*\?\/\\]", "_", sheet)[:31]
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]
            for i, col in enumerate(df.columns):
                ws.write(0, i, col, fmt_header)
                width = min(max(12, len(str(col)) + 2), 34)
                ws.set_column(i, i, width, fmt_text)
                lc = str(col).lower()
                if any(t in lc for t in ["valor", "pagado", "provision", "proyectado", "dif_pagado", "costo"]):
                    if "cantidad" not in lc and "pct" not in lc and "desv" not in lc and "horas" not in lc:
                        ws.set_column(i, i, max(width, 16), fmt_money)
                if "cantidad" in lc or "horas" in lc:
                    ws.set_column(i, i, max(width, 14), fmt_qty)
                if lc == "hc":
                    ws.set_column(i, i, max(width, 10), fmt_int)
                if "pct" in lc or "desv_pct" in lc:
                    ws.set_column(i, i, max(width, 12), fmt_pct)
            ws.autofilter(0, 0, max(len(df), 1), max(len(df.columns)-1, 0))
            ws.freeze_panes(1, 0)
    return output.getvalue()



def safe_download_button(label: str, data: bytes, file_name: str, mime: str, key: str = None):
    """Botón de descarga que no dispara rerun. Evita que Streamlit reprocesa y tumbe la app al descargar."""
    try:
        return st.download_button(
            label,
            data=data,
            file_name=file_name,
            mime=mime,
            width="stretch",
            key=key,
            on_click="ignore",
        )
    except TypeError:
        return st.download_button(
            label,
            data=data,
            file_name=file_name,
            mime=mime,
            width="stretch",
            key=key,
        )


def make_excel_ejecutivo(report: Dict[str, pd.DataFrame], extras: Dict[str, pd.DataFrame]) -> bytes:
    """Excel liviano para presentación. No incluye raw ni maestro gigante.

    El detalle completo queda dentro del paquete homologado ZIP. Esto evita caídas
    por memoria al generar/descargar Excel en Streamlit Cloud.
    """
    keep_report = [
        "Resumen_Ejecutivo_Sin_Ceros",
        "Resumen_Ejecutivo",
        "Resumen_Mes",
        "Resumen_Cargo_Homologado",
        "Indicadores_HC",
    ]
    keep_extras = [
        "Alertas",
        "Pendientes_Homologacion",
        "Homologacion_Detalle_Horas",
        "Headcount_Excluido",
    ]
    light_report = {k: report.get(k, pd.DataFrame()) for k in keep_report if k in report}
    light_extras = {k: extras.get(k, pd.DataFrame()) for k in keep_extras if k in extras}
    return make_excel(light_report, light_extras)

# ==============================
# Procesamiento orquestador
# ==============================
def process_all(inputs: Dict[str, List[Any]], umbral: float = 15.0) -> Dict[str, Any]:
    progress = st.progress(0, text="Iniciando procesamiento...")
    alerts = []
    progress.progress(5, text="Leyendo Detalle Horas y homologación final...")
    detalle_file = inputs.get("detalle", [None])[0] if inputs.get("detalle") else None
    concept_map, func_map, detalle_audit = read_detalle_horas(detalle_file)

    progress.progress(15, text="Procesando Headcount...")
    hc_full, hc_group, hc_excl, a = process_headcount(inputs.get("headcount", []), concept_map, func_map)
    alerts.extend(a)

    progress.progress(25, text="Leyendo posiciones homologadas opcionales...")
    poshom_df = None
    if inputs.get("posiciones"):
        try:
            try:
                poshom_df = read_excel_upload(inputs["posiciones"][0], sheet_name="HEAD_COUNT")
            except Exception:
                poshom_df = read_any_upload(inputs["posiciones"][0])
            alerts.append({"tipo":"Cargue", "mensaje":f"Posiciones homologadas: {len(poshom_df):,} registros leídos"})
        except Exception as e:
            alerts.append({"tipo":"Error", "mensaje":f"Error leyendo Posiciones homologadas: {e}"})

    progress.progress(35, text="Construyendo Maestro Posición → Función...")
    master, master_audit = build_position_function_master(hc_full, poshom_df)
    alerts.append({"tipo":"Homologación", "mensaje":f"Maestro Posición → Función construido con {len(master):,} llaves únicas"})

    progress.progress(40, text="Calculando HC desde posiciones homologadas por periodo...")
    hc_full, hc_group, hc_excl, a = finalize_headcount_with_master(hc_full, master, func_map)
    alerts.extend(a)

    progress.progress(45, text="Procesando pagado real CCNómina + compensatorios...")
    pagado_full, pagado_agg, a = process_pagado(inputs.get("ccnomina", []), inputs.get("compensatorios", []), concept_map, func_map, master, hc_full)
    alerts.extend(a)

    progress.progress(60, text="Procesando provisión...")
    provision_full, provision_agg, a = process_provision(inputs.get("provision", []), concept_map, func_map, master, hc_full)
    alerts.extend(a)

    progress.progress(75, text="Procesando proyección...")
    proyeccion_full, proyeccion_agg, a = process_proyeccion(inputs.get("proyeccion", []), concept_map, func_map, master, hc_full)
    alerts.extend(a)

    progress.progress(88, text="Generando comparativos y resúmenes...")
    report = aggregate_sources(pagado_agg, provision_agg, proyeccion_agg, hc_group)
    alertas_df = build_alerts(report, alerts, umbral=umbral)

    progress.progress(98, text="Preparando auditorías...")
    pendientes = pd.concat([
        pagado_full.assign(origen_base="Pagado") if pagado_full is not None and not pagado_full.empty else pd.DataFrame(),
        provision_full.assign(origen_base="Provisión") if provision_full is not None and not provision_full.empty else pd.DataFrame(),
        proyeccion_full.assign(origen_base="Proyección") if proyeccion_full is not None and not proyeccion_full.empty else pd.DataFrame(),
    ], ignore_index=True)
    if not pendientes.empty:
        pendientes = pendientes[pendientes["cargo_homologado"].isin(["", "Sin homologar"])][[c for c in ["origen_base","archivo","periodo_novedad","sap","posicion_original","funcion_codigo_final","funcion_nombre_final","area_negocio","concepto","valor_pagado","valor_provisionado","valor_proyectado"] if c in pendientes.columns]].copy()
    extras = {
        "Alertas": alertas_df,
        "Maestro_Posicion_Funcion": master_audit,
        "Homologacion_Detalle_Horas": detalle_audit,
        "Headcount_Usado": hc_group,
        "Headcount_Excluido": hc_excl,
        "Pendientes_Homologacion": pendientes,
    }
    progress.progress(100, text="Listo.")
    return {
        "report": report,
        "extras": extras,
        "raw": {
            "pagado": pagado_full, "provision": provision_full, "proyeccion": proyeccion_full,
            "headcount": hc_full,
        },
        "metrics": {
            "pagado_registros": len(pagado_full) if pagado_full is not None else 0,
            "provision_registros": len(provision_full) if provision_full is not None else 0,
            "proyeccion_registros": len(proyeccion_full) if proyeccion_full is not None else 0,
            "hc_registros": len(hc_full) if hc_full is not None else 0,
            "maestro_llaves": len(master),
        }
    }


# ==============================
# Predicción financiera - V11
# ==============================
DEFAULT_FACTORES = pd.DataFrame([
    {"concepto": "Y220", "tipo_hora": "Rec. Noc.", "factor_concepto": 0.35},
    {"concepto": "Y221", "tipo_hora": "Rec. Dom noc", "factor_concepto": 1.10},
    {"concepto": "Y300", "tipo_hora": "Hora Extra", "factor_concepto": 1.25},
    {"concepto": "Y305", "tipo_hora": "Hora Extra", "factor_concepto": 1.75},
    {"concepto": "Y310", "tipo_hora": "Hora Extra", "factor_concepto": 2.00},
    {"concepto": "Y315", "tipo_hora": "Hora Extra", "factor_concepto": 2.50},
    {"concepto": "Y350", "tipo_hora": "Compensatorio", "factor_concepto": 1.00},
    {"concepto": "YM01", "tipo_hora": "Rec. Dom", "factor_concepto": 0.75},
])

CONCEPTOS_SALARIO_BASE = {"Y050", "Y010", "Y090", "Y011", "Y051", "Y020"}
CONCEPTOS_BONOS = {"Y506", "Y610", "Y618", "Y617"}
CONCEPTOS_SALARIALES = CONCEPTOS_SALARIO_BASE | CONCEPTOS_BONOS


def period_to_year_month(period: str) -> Tuple[int, int]:
    m = re.match(r"(0[1-9]|1[0-2])\.(20\d{2})", str(period))
    if not m:
        return 0, 0
    return int(m.group(2)), int(m.group(1))


def next_period(period: str) -> str:
    y, m = period_to_year_month(period)
    if not y:
        return ""
    m += 1
    if m == 13:
        m = 1
        y += 1
    return f"{m:02d}.{y}"


def date_from_period(period: str, day: int = 1) -> pd.Timestamp:
    y, m = period_to_year_month(period)
    if not y:
        return pd.NaT
    import calendar as _cal
    last = _cal.monthrange(y, m)[1]
    return pd.Timestamp(year=y, month=m, day=min(day, last))


def suggest_calendar(period: str) -> Dict[str, int]:
    import calendar as _cal
    y, m = period_to_year_month(period)
    if not y:
        return {"dias_mes": 30, "domingos": 4, "festivos": 0}
    dias_mes = _cal.monthrange(y, m)[1]
    domingos = sum(1 for d in range(1, dias_mes + 1) if pd.Timestamp(year=y, month=m, day=d).weekday() == 6)
    festivos = 0
    try:
        import holidays
        co_holidays = holidays.country_holidays("CO", years=[y])
        festivos = sum(1 for d in range(1, dias_mes + 1) if pd.Timestamp(year=y, month=m, day=d).date() in co_holidays)
    except Exception:
        festivos = 0
    return {"dias_mes": int(dias_mes), "domingos": int(domingos), "festivos": int(festivos)}


def to_datetime_sap(s: Any) -> pd.Timestamp:
    if pd.isna(s):
        return pd.NaT
    if isinstance(s, (pd.Timestamp, datetime)):
        return pd.Timestamp(s)
    txt = str(s).strip()
    if not txt:
        return pd.NaT
    return pd.to_datetime(txt, dayfirst=True, errors="coerce")


def read_md_actual(md_file, default_jornada: float, concept_map: Dict[str, str], func_map: Dict[str, str], master: Dict[str, Dict[str, str]]) -> Tuple[pd.DataFrame, pd.DataFrame, List[Dict[str, Any]]]:
    alerts = []
    if md_file is None:
        return pd.DataFrame(), pd.DataFrame(), [{"tipo": "Predicción", "mensaje": "No se cargó MD actual; no se puede calcular costo por salario."}]
    try:
        df = read_any_upload(md_file)
        df = drop_duplicated_columns(df)
        c_sap = find_col(df, ["Nº pers.", "N° pers.", "SAP"], True)
        c_status = find_col(df, ["Status ocupación", "Status ocupacion", "Status"], False)
        c_baja = find_col(df, ["Baja"], False)
        c_div = find_col(df, ["División de personal", "Division de personal"], False)
        c_area_pers = find_col(df, ["Área de personal", "Area de personal"], False)
        c_area_nom = find_col(df, ["Área de nómina", "Area de nomina", "Texto área nómina"], False)
        c_ceco = find_col(df, ["Ce.coste", "CECO", "Ce coste"], False)
        c_pos_code = find_col(df, ["Posición", "Posicion"], False)
        c_pos_text = find_col(df, ["Posición.1", "Posicion.1", "Posición_3", "Posicion_3", "Posición                                ", "Posicion nombre", "Nombre posición"], False)
        c_func_code = find_col(df, ["Función", "Funcion"], False)
        c_func_text = find_col(df, ["Función.1", "Funcion.1", "Función_4", "Funcion_4", "Función                                 ", "Nombre función"], False)
        c_concept = find_col(df, ["CC-nómina", "CC-nomina", "CC-nómina.1", "CC-n.", "CC-n"], True)
        # Si encuentra texto de concepto en vez de código, reintentar por nombre exacto aproximado de código SAP
        # En el TXT de SAP hay dos columnas CC-nómina: código y texto; find_col suele tomar la primera.
        if c_concept and not df[c_concept].astype(str).str.upper().str.contains(r"Y\d{3}|Y\w\d", regex=True, na=False).any():
            for c in df.columns:
                if df[c].astype(str).str.upper().str.contains(r"Y\d{3}|YM01", regex=True, na=False).any():
                    c_concept = c
                    break
        c_importe = find_col(df, ["Importe", "     Importe", "Valor"], True)
        c_hsem = find_col(df, ["H sem.", "H sem", "Horas semanales", "Horas semana"], False)
        c_desde = find_col(df, ["Desde"], False)
        c_hasta = find_col(df, ["Hasta"], False)
        c_modif = find_col(df, ["Modif.el", "Modif el", "Modificado el"], False)
        out = pd.DataFrame(index=df.index)
        out["sap"] = df[c_sap].apply(clean_sap)
        out["status"] = df[c_status].apply(clean_text) if c_status else ""
        out["baja"] = df[c_baja].apply(clean_text) if c_baja else ""
        out["division"] = df[c_div].apply(clean_text) if c_div else ""
        out["area_personal"] = df[c_area_pers].apply(clean_text) if c_area_pers else ""
        out["area_nomina"] = df[c_area_nom].apply(clean_text) if c_area_nom else ""
        out["ceco"] = df[c_ceco].apply(clean_code) if c_ceco else ""
        out["posicion_codigo"] = df[c_pos_code].apply(clean_code) if c_pos_code else ""
        out["posicion_nombre"] = df[c_pos_text].apply(clean_text) if c_pos_text else ""
        out["funcion_codigo"] = df[c_func_code].apply(clean_code) if c_func_code else ""
        out["funcion_nombre"] = df[c_func_text].apply(clean_text) if c_func_text else ""
        out["concepto_salario"] = df[c_concept].apply(clean_concept)
        out["importe"] = df[c_importe].apply(parse_number)
        out["h_sem"] = df[c_hsem].apply(parse_number) if c_hsem else 0.0
        out["desde"] = df[c_desde].apply(to_datetime_sap) if c_desde else pd.NaT
        out["hasta"] = df[c_hasta].apply(to_datetime_sap) if c_hasta else pd.NaT
        out["modif_el"] = df[c_modif].apply(to_datetime_sap) if c_modif else pd.NaT
        out = out[out["sap"].ne("")].copy()
        out = out[out["concepto_salario"].isin(CONCEPTOS_SALARIALES)].copy()
        if out.empty:
            return pd.DataFrame(), pd.DataFrame(), [{"tipo": "Predicción", "mensaje": "MD actual sin conceptos salariales reconocidos para calcular salario total."}]
        # Vigencia objetivo por persona: 31.12.9999 si existe; si no, fecha máxima de Hasta.
        far_date = pd.Timestamp(year=9999, month=12, day=31)
        # Pandas no soporta year 9999 en datetime64 ns, por eso la fecha SAP 31.12.9999 puede quedar NaT. Detectar por texto original en df si aplica.
        if c_hasta:
            out["hasta_txt"] = df.loc[out.index, c_hasta].astype(str).str.strip()
            out["es_vigente_abierta"] = out["hasta_txt"].str.contains("31.12.9999|9999", regex=True, na=False)
        else:
            out["hasta_txt"] = ""
            out["es_vigente_abierta"] = False
        vig_rows = []
        for sap, g in out.groupby("sap"):
            if g["es_vigente_abierta"].any():
                target = "ABIERTA"
                gg = g[g["es_vigente_abierta"]].copy()
            else:
                max_hasta = g["hasta"].max()
                target = max_hasta
                gg = g[g["hasta"].eq(max_hasta)].copy() if pd.notna(max_hasta) else g.copy()
            gg["vigencia_usada"] = str(target)
            vig_rows.append(gg)
        vig = pd.concat(vig_rows, ignore_index=True) if vig_rows else pd.DataFrame()
        if vig.empty:
            return pd.DataFrame(), pd.DataFrame(), alerts
        vig["modif_ord"] = vig["modif_el"].fillna(pd.Timestamp("1900-01-01"))
        vig = vig.sort_values(["sap", "concepto_salario", "modif_ord", "importe"], ascending=[True, True, False, False])
        det = vig.drop_duplicates(["sap", "concepto_salario"], keep="first").copy()
        # datos dimensionales: tomar registro más reciente dentro de vigencia por persona
        dim = det.sort_values(["sap", "modif_ord"], ascending=[True, False]).drop_duplicates("sap", keep="first").copy()
        sal = det.groupby("sap", dropna=False).agg(
            salario_base=("importe", lambda s: float(det.loc[s.index][det.loc[s.index, "concepto_salario"].isin(CONCEPTOS_SALARIO_BASE)]["importe"].sum())),
            bonos_sumados=("importe", lambda s: float(det.loc[s.index][det.loc[s.index, "concepto_salario"].isin(CONCEPTOS_BONOS)]["importe"].sum())),
        ).reset_index()
        sal["salario_total"] = sal["salario_base"] + sal["bonos_sumados"]
        dim_cols = ["sap", "division", "area_personal", "area_nomina", "ceco", "posicion_codigo", "posicion_nombre", "funcion_codigo", "funcion_nombre", "h_sem", "vigencia_usada"]
        md = sal.merge(dim[dim_cols], on="sap", how="left")
        md["jornada_vigente"] = np.where(md["h_sem"].fillna(0) > 0, md["h_sem"].astype(float) * 5, float(default_jornada))
        md["jornada_vigente"] = md["jornada_vigente"].replace(0, float(default_jornada)).fillna(float(default_jornada))
        md["valor_hora"] = np.where(md["jornada_vigente"] > 0, md["salario_total"] / md["jornada_vigente"], 0)
        md["area_negocio"] = [classify_area(ceco, div, "", an, pos) for ceco, div, an, pos in zip(md["ceco"], md["division"], md["area_nomina"], md["posicion_nombre"])]
        md["manager_excluido"] = [is_manager_excl(ap, an, pos) for ap, an, pos in zip(md["area_personal"], md["area_nomina"], md["posicion_nombre"])]
        # Homologar por función; si no hay función, usar maestro Posición -> Función
        tmp = pd.DataFrame({
            "sap": md["sap"],
            "periodo_novedad": "",
            "posicion_original": md["posicion_nombre"],
            "funcion_codigo": md["funcion_codigo"],
            "funcion_nombre": md["funcion_nombre"],
            "area_negocio": md["area_negocio"],
        })
        tmp_alerts = []
        tmp = add_function_and_cargo(tmp, "MD actual", master, func_map, pd.DataFrame(), tmp_alerts, prefer_sap=False)
        md["funcion_codigo_final"] = tmp["funcion_codigo_final"].values
        md["funcion_nombre_final"] = tmp["funcion_nombre_final"].values
        md["cargo_homologado"] = tmp["cargo_homologado"].values
        md["metodo_cargo"] = tmp["metodo_cargo"].values
        alerts.append({"tipo": "Predicción", "mensaje": f"MD actual: {len(md):,} personas con salario total calculado."})
        sin_salario = md[md["salario_total"].fillna(0) <= 0]
        if len(sin_salario):
            alerts.append({"tipo": "Predicción", "mensaje": f"MD actual: {len(sin_salario):,} personas sin salario total o en cero."})
        return md, det, alerts
    except Exception as e:
        return pd.DataFrame(), pd.DataFrame(), [{"tipo": "Error", "mensaje": f"Error procesando MD actual: {e}"}]


def process_interfaces(interface_files: List[Any], concept_map: Dict[str, str], func_map: Dict[str, str], master: Dict[str, Dict[str, str]], hc_full: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, List[Dict[str, Any]]]:
    rows = []
    alerts = []
    for f in interface_files or []:
        try:
            df = read_any_upload(f)
            df = drop_duplicated_columns(df)
            # Si no reconoce encabezados, usar primeras 4 columnas: SAP, periodo/fecha pago, concepto, cantidad.
            c_sap = find_col(df, ["SAP", "Nº pers.", "N° pers.", "Nro Personal", "Número de personal"], False)
            c_period = find_col(df, ["Periodo", "Mes", "Fecha", "Fecha pago", "Mes pago"], False)
            c_concept = find_col(df, ["Concepto", "CC-n.", "CC-n", "Concepto nómina", "Concepto nomina"], False)
            c_qty = find_col(df, ["Cantidad", "Horas", "Total"], False)
            c_ceco = find_col(df, ["CECO", "Ce.coste", "Ce coste"], False)
            c_func = find_col(df, ["Función", "Funcion", "Cargo", "Posición", "Posicion"], False)
            if not all([c_sap, c_concept, c_qty]) and len(df.columns) >= 4:
                cols = list(df.columns)
                c_sap = c_sap or cols[0]
                c_period = c_period or cols[1]
                c_concept = c_concept or cols[2]
                c_qty = c_qty or cols[3]
            out = pd.DataFrame(index=df.index)
            out["fuente"] = "Interface"
            out["archivo"] = f.name
            out["sap"] = df[c_sap].apply(clean_sap) if c_sap else ""
            out["periodo_pago"] = df[c_period].apply(lambda x: parse_period_any(x, f.name)) if c_period else parse_period_any("", f.name)
            out["periodo_novedad"] = out["periodo_pago"].apply(prev_period)
            out["concepto_interface"] = df[c_concept].apply(clean_concept) if c_concept else ""
            out["concepto"] = out["concepto_interface"].map(lambda x: INTERFAZ_MAP.get(x, x))
            out["tipo_hora"] = out["concepto"].map(concept_map).fillna(out["concepto"].map(CONCEPTOS)).fillna("Sin tipo hora")
            out["cantidad_interface"] = df[c_qty].apply(parse_number) if c_qty else 0.0
            out["ceco"] = df[c_ceco].apply(clean_code) if c_ceco else ""
            func_txt = df[c_func].apply(clean_text) if c_func else pd.Series([""] * len(df))
            out["posicion_original"] = func_txt
            out["funcion_codigo"] = ""
            out["funcion_nombre"] = func_txt
            out["area_negocio"] = [classify_area(ceco, "", "", "", pos) for ceco, pos in zip(out["ceco"], func_txt)]
            out = out[out["concepto"].isin(CONCEPTOS_SET)].copy()
            out = out[out["cantidad_interface"].abs() > 0].copy()
            invalid = out["sap"].eq("").sum()
            if invalid:
                alerts.append({"tipo": "Interface", "mensaje": f"{f.name}: {invalid:,} registros sin SAP válido; se intentará homologar por función/cargo."})
            out = add_function_and_cargo(out, "Interface", master, func_map, hc_full, alerts, prefer_sap=True)
            rows.append(out)
            alerts.append({"tipo": "Cargue", "mensaje": f"Interface {f.name}: {len(out):,} registros útiles procesados"})
        except Exception as e:
            alerts.append({"tipo": "Error", "mensaje": f"Error procesando interface {getattr(f, 'name', 'archivo')}: {e}"})
    full = pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()
    if full.empty:
        return full, pd.DataFrame(), alerts
    agg = full.groupby(KEY_EXEC, dropna=False).agg(cantidad_interface=("cantidad_interface", "sum")).reset_index()
    agg = ensure_key_types(agg, KEY_EXEC)
    return full, agg, alerts


def read_cuentas(cuenta_file) -> Tuple[pd.DataFrame, List[Dict[str, Any]]]:
    alerts = []
    if cuenta_file is None:
        return pd.DataFrame(), alerts
    try:
        df = read_any_upload(cuenta_file)
        df = drop_duplicated_columns(df)
        c_con = find_col(df, ["Concepto", "CC-n.", "CC-n"], False)
        c_area = find_col(df, ["Área negocio", "Area negocio", "Area", "Tipo"], False)
        c_ceco_pref = find_col(df, ["CECO inicio", "Prefijo CECO", "CECO", "Ceco"], False)
        c_cuenta = find_col(df, ["Cuenta", "Cuenta contable", "DKON"], True)
        c_desc = find_col(df, ["Descripción", "Descripcion", "Texto cuenta"], False)
        out = pd.DataFrame(index=df.index)
        out["concepto"] = df[c_con].apply(clean_concept) if c_con else ""
        out["area_negocio"] = df[c_area].apply(clean_text) if c_area else ""
        out["ceco_prefijo"] = df[c_ceco_pref].apply(clean_code) if c_ceco_pref else ""
        out["cuenta"] = df[c_cuenta].apply(clean_code)
        out["descripcion_cuenta"] = df[c_desc].apply(clean_text) if c_desc else ""
        out = out[out["cuenta"].ne("")].copy()
        alerts.append({"tipo": "Predicción", "mensaje": f"Cuentas: {len(out):,} reglas cargadas."})
        return out, alerts
    except Exception as e:
        return pd.DataFrame(), [{"tipo": "Error", "mensaje": f"Error leyendo cuentas: {e}"}]


def assign_account(df: pd.DataFrame, cuentas: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["cuenta"] = ""
    out["descripcion_cuenta"] = ""
    if cuentas is not None and not cuentas.empty:
        # prioridad: concepto+área; concepto+prefijo; área; prefijo; concepto
        rules = cuentas.copy()
        rules["concepto"] = rules["concepto"].fillna("").astype(str)
        rules["area_negocio"] = rules["area_negocio"].fillna("").astype(str)
        rules["ceco_prefijo"] = rules["ceco_prefijo"].fillna("").astype(str)
        for idx in out.index:
            concepto = str(out.at[idx, "concepto"])
            area = str(out.at[idx, "area_negocio"])
            ceco = str(out.at[idx, "ceco"] if "ceco" in out.columns else "")
            candidates = rules.copy()
            def ok_rule(r):
                if r["concepto"] and r["concepto"] != concepto:
                    return False
                if r["area_negocio"] and r["area_negocio"] != area:
                    return False
                if r["ceco_prefijo"] and not ceco.startswith(r["ceco_prefijo"]):
                    return False
                return True
            m = candidates[candidates.apply(ok_rule, axis=1)]
            if not m.empty:
                # más específica = más campos llenos
                m = m.assign(spec=m[["concepto", "area_negocio", "ceco_prefijo"]].ne("").sum(axis=1)).sort_values("spec", ascending=False)
                out.at[idx, "cuenta"] = m.iloc[0]["cuenta"]
                out.at[idx, "descripcion_cuenta"] = m.iloc[0].get("descripcion_cuenta", "")
    # fallback por CECO / área según Modelo Financiero Nómina
    if "ceco" not in out.columns:
        out["ceco"] = ""
    empty = out["cuenta"].eq("")
    out.loc[empty & out["ceco"].astype(str).str.startswith("101"), "cuenta"] = "60"
    out.loc[empty & out["ceco"].astype(str).str.startswith("102"), "cuenta"] = "62"
    out.loc[empty & out["ceco"].astype(str).str.startswith("103"), "cuenta"] = "63"
    empty = out["cuenta"].eq("")
    out.loc[empty & out["area_negocio"].eq("Tiendas"), "cuenta"] = "60"
    out.loc[empty & out["area_negocio"].isin(["CEDI", "BDC"]), "cuenta"] = "62"
    out.loc[empty & out["area_negocio"].eq("Oficina Soporte"), "cuenta"] = "63"
    out.loc[out["descripcion_cuenta"].eq("") & out["cuenta"].eq("60"), "descripcion_cuenta"] = "Tiendas"
    out.loc[out["descripcion_cuenta"].eq("") & out["cuenta"].eq("62"), "descripcion_cuenta"] = "Logística / BDC / CEDI"
    out.loc[out["descripcion_cuenta"].eq("") & out["cuenta"].eq("63"), "descripcion_cuenta"] = "Administración / Oficina soporte"
    return out


def calendar_factor_for_concept(concepto: str, dias_mes: int, domingos: int, festivos: int, factor_manual: float) -> float:
    # El factor manual permite ajustar por calendario sin sobrecomplejizar. Se aplica a todos.
    # Los recargos dominicales/festivos quedan más sensibles si el usuario cambia factor_manual.
    return float(factor_manual)


def weighted_rate_from_history(report_exec: pd.DataFrame, months: List[str], weights: List[float]) -> pd.DataFrame:
    if report_exec is None or report_exec.empty:
        return pd.DataFrame(columns=KEY_EXEC + ["tasa_hist_horas_hc", "valor_hist_por_hora"])
    hist = report_exec[report_exec["periodo_novedad"].astype(str).isin(months)].copy()
    if hist.empty:
        return pd.DataFrame(columns=KEY_EXEC + ["tasa_hist_horas_hc", "valor_hist_por_hora"])
    wmap = {m: weights[i] if i < len(weights) else 0 for i, m in enumerate(months)}
    hist["peso"] = hist["periodo_novedad"].map(wmap).fillna(0).astype(float)
    hist["tasa_mes"] = np.where(hist["hc"].fillna(0) > 0, hist["cantidad_pagada"].fillna(0) / hist["hc"].fillna(0), 0)
    hist["valor_hora_pagado_real"] = np.where(hist["cantidad_pagada"].abs() > 0, hist["valor_pagado"] / hist["cantidad_pagada"], 0)
    gcols = [c for c in KEY_EXEC if c in hist.columns]
    out = hist.groupby(gcols, dropna=False).apply(lambda g: pd.Series({
        "tasa_hist_horas_hc": float(np.average(g["tasa_mes"], weights=np.where(g["peso"]>0, g["peso"], 0.0001))) if len(g) else 0,
        "valor_hist_por_hora": float(np.average(g["valor_hora_pagado_real"], weights=np.where(g["cantidad_pagada"].abs()>0, g["cantidad_pagada"].abs(), 0.0001))) if len(g) else 0,
    })).reset_index()
    return out


def build_prediction(res: Dict[str, Any], interface_files: List[Any], md_file, cuenta_file, periodo_prediccion: str, default_jornada: float, factores_df: pd.DataFrame, pesos: Dict[str, float], factor_cal: float, dias_mes: int, domingos: int, festivos: int) -> Dict[str, pd.DataFrame]:
    alerts = []
    report = res.get("report", {}) if res else {}
    raw = res.get("raw", {}) if res else {}
    extras = res.get("extras", {}) if res else {}
    # mapas desde procesamiento histórico
    # Releer Detalle/maestro no está disponible como dict; se puede reconstruir mínimo desde extras + raw.
    # Para robustez, usar el func_map/final ya embebido en master no es posible aquí, por eso guardaremos en res desde process_all_v11.
    func_map = res.get("func_map", {})
    concept_map = res.get("concept_map", dict(CONCEPTOS))
    master = res.get("master", {})
    hc_full = raw.get("headcount", pd.DataFrame())
    interfaz_full, interfaz_agg, a = process_interfaces(interface_files, concept_map, func_map, master, hc_full)
    alerts.extend(a)
    md, md_detalle_salario, a = read_md_actual(md_file, default_jornada, concept_map, func_map, master)
    alerts.extend(a)
    cuentas, a = read_cuentas(cuenta_file)
    alerts.extend(a)
    if factores_df is None or factores_df.empty:
        factores_df = DEFAULT_FACTORES.copy()
    factores_df = factores_df.copy()
    factores_df["concepto"] = factores_df["concepto"].apply(clean_concept)
    factores_df["factor_concepto"] = factores_df["factor_concepto"].apply(parse_number)
    prev_pred = prev_period(periodo_prediccion)
    # Históricos usados: tres meses inmediatamente anteriores al mes previo disponible
    hist_months = []
    p = prev_pred
    for _ in range(3):
        if p:
            hist_months.append(p)
            p = prev_period(p)
    # Pesos de meses: más reciente primero
    hist_weights = [0.50, 0.30, 0.20]
    exec_df = report.get("Resumen_Ejecutivo", pd.DataFrame()).copy()
    hist_rate = weighted_rate_from_history(exec_df, hist_months, hist_weights)
    # Tasa interface del mes inmediatamente anterior: interface del pago actual corresponde al mes anterior laborado.
    if interfaz_agg is not None and not interfaz_agg.empty:
        int_prev = interfaz_agg[interfaz_agg["periodo_novedad"].eq(prev_pred)].copy()
    else:
        int_prev = pd.DataFrame(columns=KEY_EXEC + ["cantidad_interface"])
    hc_hist = report.get("Resumen_Cargo_Homologado", pd.DataFrame())
    if hc_hist is not None and not hc_hist.empty and not int_prev.empty:
        hc_prev = hc_hist[hc_hist["periodo_novedad"].eq(prev_pred)][KEY_HC + ["hc"]].drop_duplicates(KEY_HC)
        int_prev = int_prev.merge(hc_prev, on=KEY_HC, how="left")
        int_prev["hc"] = int_prev["hc"].fillna(0)
        int_prev["tasa_interface_horas_hc"] = np.where(int_prev["hc"] > 0, int_prev["cantidad_interface"] / int_prev["hc"], 0)
    else:
        int_prev["tasa_interface_horas_hc"] = 0.0
    # Proyección y provisión del mes a predecir como señales
    curr_exec = exec_df[exec_df["periodo_novedad"].eq(periodo_prediccion)].copy() if exec_df is not None and not exec_df.empty else pd.DataFrame()
    if not curr_exec.empty:
        curr_exec["tasa_proy_horas_hc"] = np.where(curr_exec["hc"].fillna(0) > 0, curr_exec["cantidad_proyectada"].fillna(0) / curr_exec["hc"].fillna(0), 0)
        curr_exec["tasa_prov_horas_hc"] = np.where(curr_exec["hc"].fillna(0) > 0, curr_exec["cantidad_provisionada"].fillna(0) / curr_exec["hc"].fillna(0), 0)
        curr_signal = curr_exec[KEY_EXEC + ["cantidad_proyectada", "valor_proyectado", "cantidad_provisionada", "valor_provisionado", "tasa_proy_horas_hc", "tasa_prov_horas_hc"]].copy()
    else:
        curr_signal = pd.DataFrame(columns=KEY_EXEC + ["cantidad_proyectada", "valor_proyectado", "cantidad_provisionada", "valor_provisionado", "tasa_proy_horas_hc", "tasa_prov_horas_hc"])
    # Universo de combinaciones
    frames = []
    for df in [hist_rate, int_prev, curr_signal]:
        if df is not None and not df.empty:
            frames.append(df[KEY_EXEC].drop_duplicates())
    if frames:
        universe = pd.concat(frames, ignore_index=True).drop_duplicates()
    else:
        universe = pd.DataFrame(columns=KEY_EXEC)
    # HC actual desde MD actual, excluyendo Managers
    if md is not None and not md.empty:
        md_valid = md[~md["manager_excluido"].fillna(False)].copy()
        hc_actual = md_valid.groupby(["area_negocio", "cargo_homologado"], dropna=False).agg(
            hc_actual=("sap", "nunique"),
            valor_hora_prom=("valor_hora", "mean"),
            salario_total_prom=("salario_total", "mean"),
        ).reset_index()
    else:
        # fallback: HC del período desde comparativo si existe
        hc_fallback = hc_hist[hc_hist["periodo_novedad"].eq(periodo_prediccion)][KEY_HC + ["hc"]].copy() if hc_hist is not None and not hc_hist.empty else pd.DataFrame(columns=KEY_HC + ["hc"])
        hc_actual = hc_fallback.rename(columns={"hc": "hc_actual"})[["area_negocio", "cargo_homologado", "hc_actual"]].copy()
        hc_actual["valor_hora_prom"] = 0.0
        hc_actual["salario_total_prom"] = 0.0
    pred = universe.merge(hist_rate, on=KEY_EXEC, how="left")
    pred = pred.merge(int_prev[KEY_EXEC + ["cantidad_interface", "tasa_interface_horas_hc"]] if not int_prev.empty else pd.DataFrame(columns=KEY_EXEC + ["cantidad_interface", "tasa_interface_horas_hc"]), on=KEY_EXEC, how="left")
    pred = pred.merge(curr_signal, on=KEY_EXEC, how="left")
    pred = pred.merge(hc_actual, on=["area_negocio", "cargo_homologado"], how="left")
    pred = pred.merge(factores_df[["concepto", "factor_concepto"]], on="concepto", how="left")
    for c in ["tasa_hist_horas_hc", "tasa_interface_horas_hc", "tasa_proy_horas_hc", "tasa_prov_horas_hc", "cantidad_interface", "cantidad_proyectada", "valor_proyectado", "cantidad_provisionada", "valor_provisionado", "hc_actual", "valor_hora_prom", "factor_concepto"]:
        if c not in pred.columns:
            pred[c] = 0.0
        pred[c] = pred[c].fillna(0.0).astype(float)
    pred["factor_concepto"] = pred["factor_concepto"].replace(0, 1.0)
    wi = float(pesos.get("interface", 0.40)); wh = float(pesos.get("historico", 0.30)); wp = float(pesos.get("proyeccion", 0.20)); wv = float(pesos.get("provision", 0.10))
    total_w = max(wi + wh + wp + wv, 0.0001)
    pred["tasa_estimada_horas_hc"] = (
        pred["tasa_interface_horas_hc"] * wi +
        pred["tasa_hist_horas_hc"] * wh +
        pred["tasa_proy_horas_hc"] * wp +
        pred["tasa_prov_horas_hc"] * wv
    ) / total_w
    pred["factor_calendario"] = [calendar_factor_for_concept(c, dias_mes, domingos, festivos, factor_cal) for c in pred["concepto"]]
    pred["cantidad_estimada"] = pred["tasa_estimada_horas_hc"] * pred["hc_actual"] * pred["factor_calendario"]
    pred["valor_estimado"] = pred["cantidad_estimada"] * pred["valor_hora_prom"] * pred["factor_concepto"]
    pred["dif_estimado_vs_proyeccion"] = pred["valor_estimado"] - pred["valor_proyectado"]
    pred["dif_estimado_vs_provision"] = pred["valor_estimado"] - pred["valor_provisionado"]
    pred["periodo_prediccion"] = periodo_prediccion
    pred["periodo_pago_estimado"] = next_period(periodo_prediccion)
    pred["dias_mes"] = dias_mes
    pred["domingos"] = domingos
    pred["festivos"] = festivos
    # CECO para cuenta: no está en ejecutivo, usar vacío y fallback por área. Si quiere cuenta por CECO, se debería predecir en detalle.
    pred["ceco"] = ""
    pred = assign_account(pred, cuentas)
    pred_sin_ceros = pred[(pred["cantidad_estimada"].abs() + pred["valor_estimado"].abs() + pred["valor_proyectado"].abs() + pred["valor_provisionado"].abs()) > 0].copy()
    resumen_cuentas = pred_sin_ceros.groupby(["periodo_prediccion", "periodo_pago_estimado", "cuenta", "descripcion_cuenta", "area_negocio", "concepto", "tipo_hora"], dropna=False).agg(
        cantidad_estimada=("cantidad_estimada", "sum"),
        valor_estimado=("valor_estimado", "sum"),
        valor_proyectado=("valor_proyectado", "sum"),
        valor_provisionado=("valor_provisionado", "sum"),
    ).reset_index()
    resumen_area_cargo = pred_sin_ceros.groupby(["periodo_prediccion", "area_negocio", "cargo_homologado"], dropna=False).agg(
        hc_actual=("hc_actual", "max"),
        cantidad_estimada=("cantidad_estimada", "sum"),
        valor_estimado=("valor_estimado", "sum"),
        valor_proyectado=("valor_proyectado", "sum"),
        valor_provisionado=("valor_provisionado", "sum"),
    ).reset_index()
    # alertas predicción
    if interfaz_full.empty:
        alerts.append({"tipo": "Predicción", "mensaje": "No se cargaron interfaces; la predicción usará histórico/proyección/provisión."})
    if md is None or md.empty:
        alerts.append({"tipo": "Predicción", "mensaje": "No hay MD actual; el costo puede quedar en cero por falta de valor hora."})
    sin_hc = pred_sin_ceros[pred_sin_ceros["hc_actual"].fillna(0) == 0]
    if len(sin_hc):
        alerts.append({"tipo": "Predicción", "mensaje": f"{len(sin_hc):,} combinaciones de predicción quedaron con HC actual en cero."})
    sin_cuenta = pred_sin_ceros[pred_sin_ceros["cuenta"].fillna("").eq("")]
    if len(sin_cuenta):
        alerts.append({"tipo": "Predicción", "mensaje": f"{len(sin_cuenta):,} combinaciones de predicción sin cuenta asignada."})
    return {
        "Prediccion_Detalle": pred_sin_ceros,
        "Prediccion_Resumen_Cuenta": resumen_cuentas,
        "Prediccion_Area_Cargo": resumen_area_cargo,
        "Interfaces_Procesadas": interfaz_full,
        "MD_Valor_Hora": md if md is not None else pd.DataFrame(),
        "MD_Detalle_Salario": md_detalle_salario if md_detalle_salario is not None else pd.DataFrame(),
        "Factores_Usados": factores_df,
        "Cuentas_Usadas": cuentas if cuentas is not None else pd.DataFrame(),
        "Alertas_Prediccion": pd.DataFrame(alerts),
    }




# Compatibilidad con versiones V12/V13 que llamaban make_excel_report
def make_excel_report(report: Dict[str, pd.DataFrame], extras: Dict[str, pd.DataFrame]) -> bytes:
    return make_excel(report, extras)

def make_prediction_excel(pred: Dict[str, pd.DataFrame]) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        wb = writer.book
        fmt_header = wb.add_format({"bold": True, "bg_color": ORANGE, "font_color": "white", "border": 1})
        fmt_money = wb.add_format({"num_format": '$ #,##0', "border": 1})
        fmt_qty = wb.add_format({"num_format": '#,##0.00', "border": 1})
        fmt_text = wb.add_format({"border": 1})
        for sheet, df in pred.items():
            if df is None or df.empty:
                df = pd.DataFrame({"Mensaje": ["Sin datos"]})
            sheet_name = re.sub(r"[\[\]\:\*\?\/\\]", "_", sheet)[:31]
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]
            for i, col in enumerate(df.columns):
                ws.write(0, i, col, fmt_header)
                lc = str(col).lower()
                width = min(max(12, len(str(col)) + 2), 34)
                if any(t in lc for t in ["valor", "costo", "salario"]):
                    ws.set_column(i, i, max(width, 16), fmt_money)
                elif any(t in lc for t in ["cantidad", "horas", "tasa", "factor"]):
                    ws.set_column(i, i, max(width, 14), fmt_qty)
                else:
                    ws.set_column(i, i, width, fmt_text)
            ws.autofilter(0, 0, max(len(df), 1), max(len(df.columns)-1, 0))
            ws.freeze_panes(1, 0)
    return output.getvalue()


# Enriquecer process_all para guardar mapas de homologación en sesión
def process_all_v11(inputs: Dict[str, List[Any]], umbral: float = 15.0) -> Dict[str, Any]:
    res = process_all(inputs, umbral=umbral)
    # Reconstruir mapas para predicción desde los mismos insumos
    detalle_file = inputs.get("detalle", [None])[0] if inputs.get("detalle") else None
    concept_map, func_map, _ = read_detalle_horas(detalle_file)
    hc_full = res.get("raw", {}).get("headcount", pd.DataFrame())
    poshom_df = None
    if inputs.get("posiciones"):
        try:
            try:
                poshom_df = read_excel_upload(inputs["posiciones"][0], sheet_name="HEAD_COUNT")
            except Exception:
                poshom_df = read_any_upload(inputs["posiciones"][0])
        except Exception:
            poshom_df = None
    master, _ = build_position_function_master(hc_full, poshom_df)
    res["concept_map"] = concept_map
    res["func_map"] = func_map
    res["master"] = master
    return res


# ==============================
# Paquetes homologados - V12
# ==============================
def _df_to_csv_bytes(df: pd.DataFrame) -> bytes:
    if df is None:
        df = pd.DataFrame()
    df = df.copy()
    for c in df.columns:
        if str(c).lower().startswith("periodo"):
            df[c] = df[c].apply(normalize_period_value)
    return df.to_csv(index=False).encode("utf-8-sig")


def _csv_bytes_to_df(data: bytes) -> pd.DataFrame:
    if not data:
        return pd.DataFrame()
    df = pd.read_csv(io.BytesIO(data), dtype=str, keep_default_na=False)
    # Convertir columnas numéricas conocidas para cálculos
    numeric_tokens = [
        "valor", "cantidad", "hc", "horas", "pct", "factor", "salario", "jornada", "tasa",
        "dif_", "provision", "proyectado", "pagado", "estimado", "costo"
    ]
    for c in df.columns:
        lc = str(c).lower()
        if lc.startswith("periodo"):
            df[c] = df[c].apply(normalize_period_value)
        elif any(tok in lc for tok in numeric_tokens) and c not in ["concepto", "cuenta", "ceco", "sap"]:
            df[c] = df[c].apply(parse_number)
    return df


def maps_to_dfs(concept_map: Dict[str, str], func_map: Dict[str, str], master: Dict[str, Dict[str, str]]) -> Dict[str, pd.DataFrame]:
    concept_df = pd.DataFrame([{"concepto": k, "tipo_hora": v} for k, v in concept_map.items()])
    func_df = pd.DataFrame([{"funcion_key": k, "cargo_homologado": v} for k, v in func_map.items()])
    master_df = pd.DataFrame([{"key": k, **(v if isinstance(v, dict) else {})} for k, v in master.items()])
    return {"concept_map": concept_df, "func_map": func_df, "master": master_df}


def dfs_to_maps(dfs: Dict[str, pd.DataFrame]) -> Tuple[Dict[str, str], Dict[str, str], Dict[str, Dict[str, str]]]:
    concept_map = dict(CONCEPTOS)
    cm = dfs.get("concept_map", pd.DataFrame())
    if not cm.empty and {"concepto", "tipo_hora"}.issubset(cm.columns):
        for _, r in cm.iterrows():
            concept_map[clean_concept(r.get("concepto"))] = clean_text(r.get("tipo_hora"))
    func_map = {}
    fm = dfs.get("func_map", pd.DataFrame())
    if not fm.empty and {"funcion_key", "cargo_homologado"}.issubset(fm.columns):
        for _, r in fm.iterrows():
            func_map[clean_text(r.get("funcion_key"))] = clean_text(r.get("cargo_homologado"))
    master = {}
    md = dfs.get("master", pd.DataFrame())
    if not md.empty and "key" in md.columns:
        for _, r in md.iterrows():
            k = clean_text(r.get("key"))
            if k:
                master[k] = {
                    "funcion_codigo": clean_text(r.get("funcion_codigo", "")),
                    "funcion_nombre": clean_text(r.get("funcion_nombre", "")),
                    "posicion_codigo": clean_text(r.get("posicion_codigo", "")),
                    "posicion_nombre": clean_text(r.get("posicion_nombre", "")),
                    "origen_maestro": clean_text(r.get("origen_maestro", "")),
                }
    return concept_map, func_map, master


def make_homologated_package(res: Dict[str, Any]) -> bytes:
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("README.txt", "Paquete homologado generado por Comparativo y predicción de horas de nómina V14. Cargar este ZIP en el módulo de comparativo para análisis rápido.\n")
        for name, df in res.get("report", {}).items():
            zf.writestr(f"report/{name}.csv", _df_to_csv_bytes(df))
        for name, df in res.get("extras", {}).items():
            zf.writestr(f"extras/{name}.csv", _df_to_csv_bytes(df))
        raw_keep = ["pagado", "provision", "proyeccion", "headcount"]
        for name in raw_keep:
            df = res.get("raw", {}).get(name, pd.DataFrame())
            zf.writestr(f"raw/{name}.csv", _df_to_csv_bytes(df))
        for name, df in maps_to_dfs(res.get("concept_map", {}), res.get("func_map", {}), res.get("master", {})).items():
            zf.writestr(f"maps/{name}.csv", _df_to_csv_bytes(df))
        metrics = pd.DataFrame([res.get("metrics", {})])
        zf.writestr("metrics.csv", _df_to_csv_bytes(metrics))
    output.seek(0)
    return output.getvalue()


def load_homologated_package(uploaded_file) -> Dict[str, Any]:
    data = uploaded_file.getvalue()
    report, extras, raw, maps = {}, {}, {}, {}
    metrics = {}
    with zipfile.ZipFile(io.BytesIO(data), "r") as zf:
        for name in zf.namelist():
            if not name.endswith(".csv"):
                continue
            df = _csv_bytes_to_df(zf.read(name))
            base_name = os.path.splitext(os.path.basename(name))[0]
            if name.startswith("report/"):
                report[base_name] = df
            elif name.startswith("extras/"):
                extras[base_name] = df
            elif name.startswith("raw/"):
                raw[base_name] = df
            elif name.startswith("maps/"):
                maps[base_name] = df
            elif name == "metrics.csv":
                if not df.empty:
                    metrics = df.iloc[0].to_dict()
    concept_map, func_map, master = dfs_to_maps(maps)
    if not metrics:
        metrics = {
            "pagado_registros": len(raw.get("pagado", pd.DataFrame())),
            "provision_registros": len(raw.get("provision", pd.DataFrame())),
            "proyeccion_registros": len(raw.get("proyeccion", pd.DataFrame())),
            "hc_registros": len(raw.get("headcount", pd.DataFrame())),
            "maestro_llaves": len(master),
        }
    return {
        "report": report,
        "extras": extras,
        "raw": raw,
        "metrics": metrics,
        "concept_map": concept_map,
        "func_map": func_map,
        "master": master,
        "origen": "paquete_homologado",
    }


def get_active_result() -> Optional[Dict[str, Any]]:
    return st.session_state.get("resultado_v12")


def set_active_result(res: Dict[str, Any]) -> None:
    st.session_state["resultado_v12"] = res
    st.session_state["prediccion_v12"] = None
    st.session_state["excel_ejecutivo_bytes"] = None
    st.session_state["pred_excel_bytes"] = None


# ==============================
# UI - V12 por etapas
# ==============================
if "resultado_v12" not in st.session_state:
    st.session_state["resultado_v12"] = None
if "prediccion_v12" not in st.session_state:
    st.session_state["prediccion_v12"] = None
    st.session_state["excel_ejecutivo_bytes"] = None
    st.session_state["pred_excel_bytes"] = None
if "paquete_v12_bytes" not in st.session_state:
    st.session_state["paquete_v12_bytes"] = None
if "paquete_cargado_bytes" not in st.session_state:
    st.session_state["paquete_cargado_bytes"] = None
if "excel_ejecutivo_bytes" not in st.session_state:
    st.session_state["excel_ejecutivo_bytes"] = None
if "pred_excel_bytes" not in st.session_state:
    st.session_state["pred_excel_bytes"] = None

with st.sidebar:
    st.markdown("<div style='font-size:54px; line-height:1;'>🦜</div>", unsafe_allow_html=True)
    st.markdown("### Menú")
    page = st.radio(
        "Ir a",
        [
            "1. Preparar paquete homologado",
            "2. Cargar paquete y comparar",
            "3. Predicción financiera",
            "4. Diagnóstico y alertas",
            "5. Instructivo",
        ],
        label_visibility="collapsed",
    )
    st.divider()
    umbral = st.number_input("Umbral desviación alerta (%)", min_value=1.0, max_value=100.0, value=15.0, step=1.0)

if page == "1. Preparar paquete homologado":
    st.subheader("1. Preparar paquete homologado")
    st.markdown("""
    Este módulo hace el trabajo pesado **una sola vez**: lee los Excel originales, construye el maestro **Posición → Función**, homologa las bases y genera un ZIP liviano.

    Luego usa ese ZIP en el módulo **2. Cargar paquete y comparar** para filtrar rápido sin volver a procesar todos los archivos.
    """)
    c1, c2 = st.columns(2)
    with c1:
        detalle = st.file_uploader("Detalle Horas / Homologación (obligatorio)", type=["xlsb", "xlsx", "xlsm"], accept_multiple_files=False, key="prep_detalle")
        posiciones = st.file_uploader("Posiciones homologadas (opcional)", type=["xlsx", "xlsm"], accept_multiple_files=False, key="prep_posiciones")
        ccnomina = st.file_uploader("CCNóminas - Pagado real (cargue múltiple)", type=["xlsx", "xlsm", "xls", "txt", "csv"], accept_multiple_files=True, key="prep_cc")
        compensatorios = st.file_uploader("Compensatorios - Pagado Y350 (cargue múltiple)", type=["xls", "xlsx", "txt", "csv"], accept_multiple_files=True, key="prep_comp")
    with c2:
        headcount = st.file_uploader("Headcount mensual (cargue múltiple)", type=["xlsx", "xlsm", "xls"], accept_multiple_files=True, key="prep_hc")
        provision = st.file_uploader("Consolidado Provisión (Excel, CSV o Parquet)", type=["xlsx", "xlsm", "csv", "parquet"], accept_multiple_files=True, key="prep_prov")
        proyeccion = st.file_uploader("Consolidado Proyección", type=["xlsx", "xlsm"], accept_multiple_files=True, key="prep_proy")
    st.info("Recomendación: después de generar el paquete homologado, descárgalo y úsalo para análisis. Así los filtros y gráficas no reprocesan Excel pesados.")
    if st.button("⚙️ Procesar y generar paquete homologado", type="primary", width="stretch"):
        if not detalle:
            st.error("Debes cargar Detalle Horas / Homologación.")
        elif not headcount:
            st.error("Debes cargar al menos un Headcount para construir el maestro Posición → Función y calcular HC.")
        else:
            inputs = {
                "detalle": [detalle] if detalle else [],
                "posiciones": [posiciones] if posiciones else [],
                "ccnomina": ccnomina or [],
                "compensatorios": compensatorios or [],
                "headcount": headcount or [],
                "provision": provision or [],
                "proyeccion": proyeccion or [],
            }
            try:
                res = process_all_v11(inputs, umbral=umbral)
                set_active_result(res)
                st.session_state["paquete_v12_bytes"] = make_homologated_package(res)
                st.success("Paquete homologado generado correctamente. Descárgalo y luego úsalo en el módulo 2.")
            except Exception as e:
                st.exception(e)
    if st.session_state.get("paquete_v12_bytes"):
        safe_download_button(
            "⬇️ Descargar paquete_homologado_v14.zip",
            data=st.session_state["paquete_v12_bytes"],
            file_name=f"paquete_homologado_horas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
            mime="application/zip",
            key="download_pkg_preparado",
        )
    if get_active_result():
        m = get_active_result().get("metrics", {})
        st.subheader("Control de registros procesados")
        cols = st.columns(5)
        cols[0].metric("Pagado", format_int(m.get("pagado_registros", 0)))
        cols[1].metric("Provisión", format_int(m.get("provision_registros", 0)))
        cols[2].metric("Proyección", format_int(m.get("proyeccion_registros", 0)))
        cols[3].metric("Headcount", format_int(m.get("hc_registros", 0)))
        cols[4].metric("Maestro", format_int(m.get("maestro_llaves", 0)))

elif page == "2. Cargar paquete y comparar":
    st.subheader("2. Cargar paquete y comparar")
    paquete = st.file_uploader("Cargar paquete homologado V14 (.zip)", type=["zip"], accept_multiple_files=False, key="load_pkg")
    if paquete is not None:
        if st.button("📦 Cargar paquete homologado", type="primary"):
            try:
                res = load_homologated_package(paquete)
                st.session_state["paquete_cargado_bytes"] = paquete.getvalue()
                set_active_result(res)
                st.success("Paquete cargado. Ya puedes filtrar y descargar resultados sin reprocesar los Excel originales.")
            except Exception as e:
                st.exception(e)
    res = get_active_result()
    if not res:
        st.warning("Primero genera o carga un paquete homologado.")
    else:
        report = res.get("report", {})
        resumen_mes = report.get("Resumen_Mes", pd.DataFrame())
        total_pagado = resumen_mes["valor_pagado"].sum() if not resumen_mes.empty and "valor_pagado" in resumen_mes.columns else 0
        total_prov = resumen_mes["valor_provisionado"].sum() if not resumen_mes.empty and "valor_provisionado" in resumen_mes.columns else 0
        total_proy = resumen_mes["valor_proyectado"].sum() if not resumen_mes.empty and "valor_proyectado" in resumen_mes.columns else 0
        cols = st.columns(4)
        cols[0].metric("Pagado", format_money(total_pagado))
        cols[1].metric("Provisión", format_money(total_prov))
        cols[2].metric("Proyección", format_money(total_proy))
        cols[3].metric("Dif. pagado vs provisión", format_money(total_pagado - total_prov))
        base = prepare_report_base(report.get("Resumen_Ejecutivo_Sin_Ceros", pd.DataFrame()))
        if base.empty:
            st.warning("No hay datos en Resumen Ejecutivo. Revisa alertas y paquete generado.")
        else:
            with st.form("filtros_comparativo_v12"):
                st.markdown("#### Filtros — vacío = todos")
                fcols = st.columns(5)
                meses = safe_unique_options(base, "periodo_novedad", sort_period=True)
                areas = safe_unique_options(base, "area_negocio")
                cargos = safe_unique_options(base, "cargo_homologado")
                conceptos = safe_unique_options(base, "concepto")
                tipos = safe_unique_options(base, "tipo_hora")
                sel_m = fcols[0].multiselect("Mes novedad", meses, default=[])
                sel_a = fcols[1].multiselect("Área negocio", areas, default=[])
                sel_c = fcols[2].multiselect("Cargo homologado", cargos, default=[])
                sel_con = fcols[3].multiselect("Concepto", conceptos, default=[])
                sel_t = fcols[4].multiselect("Tipo hora", tipos, default=[])
                aplicar = st.form_submit_button("Aplicar filtros", type="primary")
            with st.spinner("Aplicando filtros en modo seguro..."):
                filt = filter_df(base, {
                    "periodo_novedad": sel_m,
                    "area_negocio": sel_a,
                    "cargo_homologado": sel_c,
                    "concepto": sel_con,
                    "tipo_hora": sel_t,
                })
            st.markdown("### Resumen ejecutivo filtrado")
            if filt.empty:
                st.info("No hay registros para los filtros seleccionados. Limpia algún filtro o selecciona otra combinación.")
            else:
                total_f_pagado = filt["valor_pagado"].sum() if "valor_pagado" in filt.columns else 0
                total_f_prov = filt["valor_provisionado"].sum() if "valor_provisionado" in filt.columns else 0
                total_f_proy = filt["valor_proyectado"].sum() if "valor_proyectado" in filt.columns else 0
                m1, m2, m3, m4 = st.columns(4)
                m1.metric("Pagado filtrado", format_money(total_f_pagado))
                m2.metric("Provisión filtrada", format_money(total_f_prov))
                m3.metric("Proyección filtrada", format_money(total_f_proy))
                m4.metric("Dif. vs provisión", format_money(total_f_pagado-total_f_prov))
                display_df(filt)
            tab1, tab2, tab3, tab4 = st.tabs(["Resumen por mes", "Resumen por cargo", "Indicadores HC", "Descargar"])
            with tab1:
                display_df(report.get("Resumen_Mes", pd.DataFrame()))
            with tab2:
                display_df(report.get("Resumen_Cargo_Homologado", pd.DataFrame()))
            with tab3:
                display_df(report.get("Indicadores_HC", pd.DataFrame()))
            with tab4:
                st.info("Para evitar caídas por memoria, el Excel descargable es ejecutivo/liviano. El detalle completo queda en el paquete homologado ZIP.")
                if st.button("📄 Generar Excel ejecutivo", key="btn_generar_excel_ejecutivo"):
                    with st.spinner("Generando Excel ejecutivo liviano..."):
                        try:
                            st.session_state["excel_ejecutivo_bytes"] = make_excel_ejecutivo(report, res.get("extras", {}))
                            st.success("Excel ejecutivo generado. Ya puedes descargarlo.")
                        except Exception as e:
                            st.error("No fue posible generar el Excel ejecutivo.")
                            st.caption(f"Detalle técnico: {type(e).__name__}: {e}")
                if st.session_state.get("excel_ejecutivo_bytes"):
                    safe_download_button(
                        "⬇️ Descargar Excel ejecutivo",
                        st.session_state["excel_ejecutivo_bytes"],
                        "comparativo_horas_resumen_ejecutivo.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="download_excel_ejecutivo",
                    )
                pkg_bytes = st.session_state.get("paquete_v12_bytes") or st.session_state.get("paquete_cargado_bytes")
                if pkg_bytes:
                    safe_download_button(
                        "⬇️ Descargar paquete homologado completo",
                        pkg_bytes,
                        "paquete_homologado_v14.zip",
                        "application/zip",
                        key="download_pkg_modulo2",
                    )

elif page == "3. Predicción financiera":
    st.subheader("3. Predicción financiera del mes en curso")
    res = get_active_result()
    if not res:
        st.warning("Primero carga o genera un paquete homologado en los módulos 1 o 2.")
    else:
        st.markdown("La predicción usa el comparativo histórico homologado + interfaces regionales + MD actual + factores + cuentas.")
        c1, c2 = st.columns(2)
        with c1:
            periodo_pred = st.text_input("Mes a predecir (MM.AAAA)", value="06.2026")
            interfaces = st.file_uploader("Interfaces regionales del mes actual", type=["xlsx", "xlsm", "xls", "txt", "csv"], accept_multiple_files=True)
            md_actual = st.file_uploader("MD actual TXT/Excel para salario total y valor hora", type=["txt", "xlsx", "xlsm", "xls", "csv"], accept_multiple_files=False)
            jornada = st.number_input("Jornada mensual por defecto", value=220.0, step=1.0)
        with c2:
            sug = suggest_calendar(periodo_pred)
            dias_mes = st.number_input("Días calendario del mes", value=float(sug["dias_mes"]), step=1.0)
            domingos = st.number_input("Domingos del mes", value=float(sug["domingos"]), step=1.0)
            festivos = st.number_input("Festivos del mes", value=float(sug["festivos"]), step=1.0)
            cuentas_file = st.file_uploader("Tabla de cuentas contables (opcional)", type=["xlsx", "xlsm", "csv"], accept_multiple_files=False)
        st.markdown("#### Factores por concepto")
        factores_df = st.data_editor(DEFAULT_FACTORES, num_rows="dynamic", width="stretch")
        st.markdown("#### Pesos de predicción")
        pcols = st.columns(4)
        peso_interfaz = pcols[0].number_input("Peso interface reciente", value=0.40, min_value=0.0, max_value=1.0, step=0.05)
        peso_ultimo = pcols[1].number_input("Peso último pago real", value=0.30, min_value=0.0, max_value=1.0, step=0.05)
        peso_hist = pcols[2].number_input("Peso promedio histórico", value=0.20, min_value=0.0, max_value=1.0, step=0.05)
        peso_plan = pcols[3].number_input("Peso proyección/provisión", value=0.10, min_value=0.0, max_value=1.0, step=0.05)
        if st.button("🔮 Generar predicción", type="primary", width="stretch"):
            try:
                pred = predict_current_month(
                    res,
                    periodo_pred=periodo_pred,
                    interface_files=interfaces or [],
                    md_file=md_actual,
                    cuentas_file=cuentas_file,
                    factores_df=factores_df,
                    default_jornada=jornada,
                    pesos={"interface": peso_interfaz, "ultimo": peso_ultimo, "historico": peso_hist, "plan": peso_plan},
                    calendario={"dias_mes": dias_mes, "domingos": domingos, "festivos": festivos},
                )
                st.session_state["prediccion_v12"] = pred
                st.session_state["pred_excel_bytes"] = None
                st.success("Predicción generada.")
            except Exception as e:
                st.exception(e)
        pred = st.session_state.get("prediccion_v12")
        if pred:
            pred_det = pred.get("Prediccion_Detalle", pd.DataFrame())
            total_est = pred_det["valor_estimado"].sum() if not pred_det.empty and "valor_estimado" in pred_det.columns else 0
            total_q = pred_det["cantidad_estimada"].sum() if not pred_det.empty and "cantidad_estimada" in pred_det.columns else 0
            k1, k2, k3 = st.columns(3)
            k1.metric("Cantidad estimada", format_qty(total_q))
            k2.metric("Valor estimado", format_money(total_est))
            k3.metric("Pago estimado", next_period(periodo_pred))
            tabs = st.tabs(["Detalle", "Por cuenta", "Área/Cargo", "Alertas", "Descargar"])
            with tabs[0]: display_df(pred_det)
            with tabs[1]: display_df(pred.get("Prediccion_Resumen_Cuenta", pd.DataFrame()))
            with tabs[2]: display_df(pred.get("Prediccion_Area_Cargo", pd.DataFrame()))
            with tabs[3]: display_df(pred.get("Alertas_Prediccion", pd.DataFrame()))
            with tabs[4]:
                if st.button("📄 Generar Excel predicción", key="btn_generar_excel_pred"):
                    with st.spinner("Generando Excel de predicción..."):
                        try:
                            st.session_state["pred_excel_bytes"] = make_prediction_excel(pred)
                            st.success("Excel de predicción generado. Ya puedes descargarlo.")
                        except Exception as e:
                            st.error("No fue posible generar el Excel de predicción.")
                            st.caption(f"Detalle técnico: {type(e).__name__}: {e}")
                if st.session_state.get("pred_excel_bytes"):
                    safe_download_button(
                        "⬇️ Descargar Excel predicción",
                        st.session_state["pred_excel_bytes"],
                        "prediccion_financiera_horas.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="download_excel_pred",
                    )

elif page == "4. Diagnóstico y alertas":
    st.subheader("4. Diagnóstico y alertas")
    res = get_active_result()
    if not res:
        st.warning("Primero procesa o carga un paquete homologado.")
    else:
        extras = res.get("extras", {})
        tabs = st.tabs(["Alertas", "Maestro posición→función", "Pendientes homologación", "Headcount usado", "Headcount excluido"])
        with tabs[0]: display_df(extras.get("Alertas", pd.DataFrame()))
        with tabs[1]: display_df(extras.get("Maestro_Posicion_Funcion", pd.DataFrame()))
        with tabs[2]: display_df(extras.get("Pendientes_Homologacion", pd.DataFrame()))
        with tabs[3]: display_df(extras.get("Headcount_Usado", pd.DataFrame()))
        with tabs[4]: display_df(extras.get("Headcount_Excluido", pd.DataFrame()))

elif page == "5. Instructivo":
    st.subheader("5. Instructivo de uso")
    st.markdown("""
    ## Objetivo
    El aplicativo compara históricamente lo **pagado** contra lo **provisionado** y lo **proyectado**, y luego usa ese histórico para generar una **predicción financiera de horas** alineada con el Modelo Financiero Nómina.

    ## Arquitectura V12 por etapas
    1. **Preparar paquete homologado:** carga los Excel originales y genera un ZIP homologado. Este paso es pesado, pero se ejecuta una sola vez por corte.
    2. **Cargar paquete y comparar:** carga el ZIP homologado y permite filtrar rápido, descargar Excel y revisar alertas sin reprocesar archivos originales.
    3. **Predicción financiera:** usa el paquete homologado + interfaces + MD actual + factores + cuentas.

    ## Regla de mes vencido
    CCNómina y compensatorios son pago real, pero se comparan por mes de novedad:
    - Pago 02.2026 corresponde a novedad 01.2026.
    - Pago 03.2026 corresponde a novedad 02.2026.

    ## Homologación
    Primero se construye un maestro **Posición → Función** desde Posiciones Homologadas o Headcount. Luego todas las fuentes pasan por:
    **Posición/Cargo original → Función → Cargo homologado**.

    ## Headcount
    El HC se agrupa por **Mes + Área negocio + Cargo homologado**, excluyendo Manager I, II, III y IV porque no aplican para estos conceptos.

    ## Predicción
    La predicción estima el mes actual que se pagará el mes siguiente. Usa histórico, interfaces recientes, proyección/provisión, HC actual, calendario y MD actual para calcular valor hora.

    ## Costo estimado
    **Valor hora = salario total vigente / jornada vigente**  
    **Costo = cantidad estimada × valor hora × factor concepto**

    ## Salidas principales
    - Comparativo histórico.
    - Resumen ejecutivo.
    - Indicadores HC.
    - Predicción detalle.
    - Predicción por cuenta.
    - Alertas y pendientes de homologación.
    """)

st.markdown("<div class='small-note'>Creado por Andrés Huérfano Dávila - Nómina JMC</div>", unsafe_allow_html=True)
